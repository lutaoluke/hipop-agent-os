"""Tool support helpers extracted from agent.py for WS-169.

These are implementation helpers for tool bodies and live ERP/logistics access.
`agent.py` re-exports them so existing tests and tools_impl dynamic patch points
continue to address `hipop.server.agent.X`.
"""
import json
import os
from typing import Any, Dict, Optional

# T03 门：SKU 销量问答必须走实时取数路径，不能以 wf2_sku 快照 as_of_date 是否新鲜为由绕过。
# injectable：测试注入 mock fn；生产使用 _erp_sku_stats_live 默认实现。
# 签名：(sku: str, nation_id: int, token: str | None) -> dict
#   ok=True:  {"sales_30d": int|None, "history_total": int|None, "fetched_at": str, "source": str}
#   ok=False: {"error": str, "message": str}
_sku_sales_live_fn: Optional[Any] = None


def _erp_sku_stats_live(sku: str, nation_id: int, token) -> dict:
    """ERP /product-order-statistics 实时拉单 SKU 30d 销量。token=None 时直接返回不可用。"""
    if token is None:
        return {"ok": False, "error": "no_erp_token",
                "message": "ERP 凭据不可用（未配置或登录失败），无法实时确认 SKU 销量"}
    import datetime as _dt
    import re as _re
    try:
        from workflows.wf0_logistics import erp_get as _erp_get
    except ImportError:
        return {"ok": False, "error": "erp_import_error",
                "message": "无法加载 ERP 模块"}
    today = _dt.date.today()
    start = today - _dt.timedelta(days=30)

    def _fmt(d):
        return f"{d.year}-{d.month}-{d.day}"

    params = {
        "nation_id": nation_id,
        "platform_id": 2,
        "keyword": sku,
        "keyword_type": 1,
        "ordered_time_section[]": [_fmt(start), _fmt(today)],
        "page": 1,
        "limit": 50,
    }
    try:
        resp = _erp_get("/product-order-statistics", params, token)
    except Exception as e:
        return {"ok": False, "error": f"erp_fetch: {type(e).__name__}",
                "message": str(e)[:200]}
    if resp.get("code") != 200:
        return {"ok": False, "error": "erp_api_error",
                "message": str(resp.get("msg") or "")[:200]}
    items = resp.get("data") or []
    sku_up = sku.upper()
    match = None
    for it in items:
        sku_obj = it.get("sku") or {}
        for psk in (sku_obj.get("platform_sku_ids") or []):
            if (psk.get("platform_sku_id") or "").upper() == sku_up:
                match = it
                break
        if not match and (it.get("sku_id") or "").upper() == sku_up:
            match = it
        if match:
            break
    if not match:
        return {"ok": False, "error": "sku_not_found_in_erp",
                "message": f"SKU {sku} 在 ERP 30d 窗口内无记录（API 可能不支持 keyword 过滤或该 SKU 无近期订单）"}

    _NATION_TO_COUNTRY = {1: "SA", 2: "AE"}
    country_code = _NATION_TO_COUNTRY.get(nation_id)
    if not country_code:
        return {"ok": False, "error": f"unknown_nation_id_{nation_id}",
                "message": f"nation_id={nation_id} 未知，无法确定目标国家前缀，拒绝返回其他国家数字"}

    def _parse_country(val, country):
        pat = _re.compile(rf"{country}\s*[:：]\s*(\d+)")
        for s in (val if isinstance(val, list) else [str(val or "")]):
            m = pat.match(str(s).strip())
            if m:
                return int(m.group(1))
        return None

    sales_30d = _parse_country(match.get("sales_count"), country_code)
    fetched_at = _dt.datetime.utcnow().isoformat() + "Z"
    return {
        "ok": True,
        "sku": sku,
        "sales_30d": sales_30d,
        "history_total": None,
        "fetched_at": fetched_at,
        "source": "ERP /product-order-statistics (realtime)",
        "window_days": 30,
    }




def _normalize_replenishment_rows(rows: list) -> list:
    out = []
    for row in rows or []:
        r = dict(row)
        if r.get("trend") == "急速下降":
            r["urgency_level"] = "high"
        elif r.get("trend") in ("下降", "加速增长"):
            r["urgency_level"] = "mid"
        else:
            r["urgency_level"] = "low"
        r["qty"] = r.get("weekly_total_replenish") or 0
        try:
            r["trigger_reasons_list"] = json.loads(r.get("trigger_reasons") or "[]")
        except Exception:
            r["trigger_reasons_list"] = []
        out.append(r)
    return out














def _write_xlsx_and_return(rows: list, name_prefix: str, filter_desc: str = "",
                             cols: list = None, headers: list = None) -> Dict:
    """统一 xlsx 写盘 + 返下载链接 helper。

    cols    决定取值用的 row dict key（顺序即列顺序）。
    headers 可选，写表头时用的展示名（与 cols 一一对应，长度须相等）；
            缺省退回 cols 自身（向后兼容老调用）。
    """
    import os, time
    from datetime import datetime
    from openpyxl import Workbook

    export_dir = os.path.expanduser("~/hipop/exports")
    os.makedirs(export_dir, exist_ok=True)

    if not rows:
        return {"ok": True, "row_count": 0,
                "message": "查询无数据 — 不生成空文件"}

    if not cols:
        cols = list(rows[0].keys()) if isinstance(rows[0], dict) else []

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"{name_prefix}_{ts}.xlsx"
    fpath = os.path.join(export_dir, fname)

    if not headers or len(headers) != len(cols):
        headers = cols

    wb = Workbook()
    ws = wb.active
    ws.title = name_prefix[:30]
    ws.append(headers)
    for r in rows:
        if isinstance(r, dict):
            ws.append([r.get(c) for c in cols])
        else:
            ws.append(list(r))
    # 第一行加粗
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    # 自适应列宽（粗略）
    for col_idx, c in enumerate(cols, 1):
        max_len = max((len(str(r.get(c) if isinstance(r, dict) else "")) for r in rows[:50]),
                      default=10)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(len(c), max_len) + 2, 40)
    wb.save(fpath)

    download_url = f"/api/download/{fname}"
    return {
        "ok": True,
        "row_count": len(rows),
        "download_url": download_url,
        "filename": fname,
        "filter_desc": filter_desc,
        "message": f"已生成 {len(rows)} 行 xlsx，下载: {download_url}",
        "file_path": fpath,  # 本地路径方便排查
    }










def _erp_token_or_error(tid: int):
    """共用：拿 per-tenant ERP token。没凭据返 'no_creds'；登录失败返 'login_failed'."""
    from . import _erp_auth
    creds = _erp_auth._get_creds(tid)
    if not creds:
        return None, {"ok": False, "error": "no_erp_credentials",
                       "message": f"tenant {tid} 没配 ERP 账号 — 请去 /onboarding 配 dbuyerp"}
    token = _erp_auth.get_erp_token_for_tenant(tid)
    if not token:
        return None, {"ok": False, "error": "erp_login_failed",
                       "message": f"ERP 凭据存在但 playwright 登录失败 "
                                   "（dbuyerp 可能在风控同账号短时间多次登），稍后重试"}
    return token, None


def _patch_wls_token(token: str):
    """共用：monkey-patch wls + wf0.get_erp_token 闭包，跟 wf3_logistics_v2 同套路"""
    import sys as _sys
    _hipop_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _hipop_dir not in _sys.path:
        _sys.path.insert(0, _hipop_dir)
    from workflows import wf0_logistics as _wf0
    from workflows import wf_logistics_status as _wls
    _orig = _wf0.get_erp_token
    _wf0.get_erp_token = lambda: token
    _wls.get_erp_token = lambda: token
    return _wf0, _wls, _orig


def _fetch_logistics_nodes(forwarder: str, tracking_no: str) -> dict:
    """playwright 实时抓物流站节点（安时达/阳光/义特/飞坦）。返回 {nodes, note}."""
    if not forwarder or not tracking_no:
        return {"nodes": [], "note": "缺 forwarder 或 tracking_no"}
    try:
        from playwright.sync_api import sync_playwright
        from workflows.wf_logistics_status import query_tracking
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            try:
                r = query_tracking(page, forwarder, tracking_no)
            finally:
                browser.close()
            return r
    except Exception as e:
        return {"nodes": [], "note": f"playwright_error: {type(e).__name__}: {str(e)[:100]}"}


def _physical_tracking_url(forwarder: str, tracking_no: str) -> str:
    """根据 forwarder 拼物流站直链。"""
    if not forwarder or not tracking_no:
        return ""
    from workflows.wf_logistics_status import LOGISTICS_URLS
    base = LOGISTICS_URLS.get(forwarder, "")
    if not base:
        return f"(无直链, forwarder={forwarder})"
    sep = "&" if "?" in base else "?"
    return f"{base}{sep}msg={tracking_no}" if "tracking" in base.lower() else f"{base}{sep}no={tracking_no}"


def _utc_now_iso() -> str:
    import datetime as _dt
    return _dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
