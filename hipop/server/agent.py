"""
LLM Coordinator - 把 hipop 的 5 个 handler + 2 个新工具包装成 Anthropic tool-use API。

7 个 tool:
  1. query_sku                   - 查 SKU 健康（来自 wf2/wf3/wf5/wf6）
  2. query_order                 - 查货单告警 + 涉及 SKU
  3. update_alert_status         - 反馈货单状态（已确认丢货 / 已约仓 / ...）
  4. scope_overview              - 店铺概览（指定国家+平台）
  5. compute_replenishment       - 列出当前店铺的补货建议
  6. compute_air_freight_roi     - 海运 vs 空运 ROI 估算（基于 SKU 利润 + 销量）
  7. data_health_check           - 数据新鲜度检查（最新 imported_at / updated_at）

每次 tool 调用都会写入 agent_actions 表 (action_type='execute')，并把 references_json
回传给前端用于"📎 出处"展示。
"""
import os, sys, json, sqlite3, time, contextvars, re
from typing import List, Dict, Any, Optional

# ── chat 当前请求 context（tenant_id + scope）─────────────
# 由 chat() 入口 set，所有 tool 函数同线程读
_chat_tenant: contextvars.ContextVar[int] = contextvars.ContextVar("chat_tenant", default=1)
_chat_scope: contextvars.ContextVar[dict] = contextvars.ContextVar("chat_scope", default={})
_chat_question: contextvars.ContextVar[str] = contextvars.ContextVar("chat_question", default="")
_last_replenishment_stock_status: contextvars.ContextVar[Optional[dict]] = contextvars.ContextVar(
    "last_replenishment_stock_status", default=None
)
_last_sku_rate_stats: contextvars.ContextVar[Optional[list]] = contextvars.ContextVar(
    "last_sku_rate_stats", default=None
)
# WS-145 肯定执行意图门:chat() 入口按本轮句式语气求出的门决策。
# _exec_tool 据此拒绝「非执行语气下偷偷 run_workflow」（LLM 不许绕）。
_chat_intent: contextvars.ContextVar[Optional[Any]] = contextvars.ContextVar(
    "chat_intent", default=None
)


def _get_tenant() -> int:
    return _chat_tenant.get() or 1


def _resolve_entity_alias(store_code: str) -> Optional[str]:
    """把工作台顶部 dropdown 的 store code（KSA/UAE）转成本租户的 entity_alias。
    按 (tenant_id, country) 查 sales_entities 表。
    KSA → SA, UAE → AE
    """
    from . import data as _d
    tid = _get_tenant()
    country = {"KSA": "SA", "UAE": "AE", "SA": "SA", "AE": "AE"}.get(store_code.upper())
    if not country:
        return None
    rows = _d._fetch(
        "SELECT alias FROM sales_entities WHERE tenant_id=? AND country=? AND active=1 LIMIT 1",
        (tid, country),
    )
    return rows[0]["alias"] if rows else None

# 让 hipop/scripts/* import
HIPOP_ROOT = os.path.dirname(os.path.dirname(__file__))
PROJECT_ROOT = os.path.dirname(HIPOP_ROOT)
sys.path.insert(0, HIPOP_ROOT)
sys.path.insert(0, PROJECT_ROOT)

from . import data as _data
from ._workflow_reply import _workflow_receipt_reply
from ._inventory_constraint_rule import handle_inventory_constraint_rule_chat as _handle_icr_chat

import anthropic
from . import _auth
from .tools_registry import load_tools_from_yaml


def _get_client():
    return _auth.get_client()


# ── 工具定义（Anthropic tool schema）────────────────────
TOOLS = load_tools_from_yaml()


# ── 工具实现（v2 列存：按 tenant_id + entity_alias 过滤）──

# T03 门：SKU 销量问答必须走实时取数路径，不能以 wf2_sku 快照 as_of_date 是否新鲜为由绕过。
# injectable：测试注入 mock fn；生产使用 _erp_sku_stats_live 默认实现。
# 签名：(sku: str, nation_id: int, token: str | None) -> dict
#   ok=True:  {"sales_30d": int|None, "history_total": int|None, "fetched_at": str, "source": str}
#   ok=False: {"error": str, "message": str}
_sku_sales_live_fn: Optional[Any] = None


def _erp_sku_stats_live(sku: str, nation_id: int, token) -> dict:
    """ERP /product-order-statistics 实时拉单 SKU 30d 销量。token=None 时直接返回不可用。"""
    if token is None:
        return {"ok": False, "error": "no_erp_token",
                "message": "ERP 凭据不可用（未配置或登录失败），无法实时确认 SKU 销量"}
    import datetime as _dt
    import re as _re
    try:
        from workflows.wf0_logistics import erp_get as _erp_get
    except ImportError:
        return {"ok": False, "error": "erp_import_error",
                "message": "无法加载 ERP 模块"}
    today = _dt.date.today()
    start = today - _dt.timedelta(days=30)

    def _fmt(d):
        return f"{d.year}-{d.month}-{d.day}"

    params = {
        "nation_id": nation_id,
        "platform_id": 2,
        "keyword": sku,
        "keyword_type": 1,
        "ordered_time_section[]": [_fmt(start), _fmt(today)],
        "page": 1,
        "limit": 50,
    }
    try:
        resp = _erp_get("/product-order-statistics", params, token)
    except Exception as e:
        return {"ok": False, "error": f"erp_fetch: {type(e).__name__}",
                "message": str(e)[:200]}
    if resp.get("code") != 200:
        return {"ok": False, "error": "erp_api_error",
                "message": str(resp.get("msg") or "")[:200]}
    items = resp.get("data") or []
    sku_up = sku.upper()
    match = None
    for it in items:
        sku_obj = it.get("sku") or {}
        for psk in (sku_obj.get("platform_sku_ids") or []):
            if (psk.get("platform_sku_id") or "").upper() == sku_up:
                match = it
                break
        if not match and (it.get("sku_id") or "").upper() == sku_up:
            match = it
        if match:
            break
    if not match:
        return {"ok": False, "error": "sku_not_found_in_erp",
                "message": f"SKU {sku} 在 ERP 30d 窗口内无记录（API 可能不支持 keyword 过滤或该 SKU 无近期订单）"}

    _NATION_TO_COUNTRY = {1: "SA", 2: "AE"}
    country_code = _NATION_TO_COUNTRY.get(nation_id)
    if not country_code:
        return {"ok": False, "error": f"unknown_nation_id_{nation_id}",
                "message": f"nation_id={nation_id} 未知，无法确定目标国家前缀，拒绝返回其他国家数字"}

    def _parse_country(val, country):
        pat = _re.compile(rf"{country}\s*[:：]\s*(\d+)")
        for s in (val if isinstance(val, list) else [str(val or "")]):
            m = pat.match(str(s).strip())
            if m:
                return int(m.group(1))
        return None

    sales_30d = _parse_country(match.get("sales_count"), country_code)
    fetched_at = _dt.datetime.utcnow().isoformat() + "Z"
    return {
        "ok": True,
        "sku": sku,
        "sales_30d": sales_30d,
        "history_total": None,
        "fetched_at": fetched_at,
        "source": "ERP /product-order-statistics (realtime)",
        "window_days": 30,
    }




def _normalize_replenishment_rows(rows: list) -> list:
    out = []
    for row in rows or []:
        r = dict(row)
        if r.get("trend") == "急速下降":
            r["urgency_level"] = "high"
        elif r.get("trend") in ("下降", "加速增长"):
            r["urgency_level"] = "mid"
        else:
            r["urgency_level"] = "low"
        r["qty"] = r.get("weekly_total_replenish") or 0
        try:
            r["trigger_reasons_list"] = json.loads(r.get("trigger_reasons") or "[]")
        except Exception:
            r["trigger_reasons_list"] = []
        out.append(r)
    return out














def _write_xlsx_and_return(rows: list, name_prefix: str, filter_desc: str = "",
                             cols: list = None, headers: list = None) -> Dict:
    """统一 xlsx 写盘 + 返下载链接 helper。

    cols    决定取值用的 row dict key（顺序即列顺序）。
    headers 可选，写表头时用的展示名（与 cols 一一对应，长度须相等）；
            缺省退回 cols 自身（向后兼容老调用）。
    """
    import os, time
    from datetime import datetime
    from openpyxl import Workbook

    export_dir = os.path.expanduser("~/hipop/exports")
    os.makedirs(export_dir, exist_ok=True)

    if not rows:
        return {"ok": True, "row_count": 0,
                "message": "查询无数据 — 不生成空文件"}

    if not cols:
        cols = list(rows[0].keys()) if isinstance(rows[0], dict) else []

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"{name_prefix}_{ts}.xlsx"
    fpath = os.path.join(export_dir, fname)

    if not headers or len(headers) != len(cols):
        headers = cols

    wb = Workbook()
    ws = wb.active
    ws.title = name_prefix[:30]
    ws.append(headers)
    for r in rows:
        if isinstance(r, dict):
            ws.append([r.get(c) for c in cols])
        else:
            ws.append(list(r))
    # 第一行加粗
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    # 自适应列宽（粗略）
    for col_idx, c in enumerate(cols, 1):
        max_len = max((len(str(r.get(c) if isinstance(r, dict) else "")) for r in rows[:50]),
                      default=10)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(len(c), max_len) + 2, 40)
    wb.save(fpath)

    download_url = f"/api/download/{fname}"
    return {
        "ok": True,
        "row_count": len(rows),
        "download_url": download_url,
        "filename": fname,
        "filter_desc": filter_desc,
        "message": f"已生成 {len(rows)} 行 xlsx，下载: {download_url}",
        "file_path": fpath,  # 本地路径方便排查
    }










def _erp_token_or_error(tid: int):
    """共用：拿 per-tenant ERP token。没凭据返 'no_creds'；登录失败返 'login_failed'."""
    from . import _erp_auth
    creds = _erp_auth._get_creds(tid)
    if not creds:
        return None, {"ok": False, "error": "no_erp_credentials",
                       "message": f"tenant {tid} 没配 ERP 账号 — 请去 /onboarding 配 dbuyerp"}
    token = _erp_auth.get_erp_token_for_tenant(tid)
    if not token:
        return None, {"ok": False, "error": "erp_login_failed",
                       "message": f"ERP 凭据存在但 playwright 登录失败 "
                                   "（dbuyerp 可能在风控同账号短时间多次登），稍后重试"}
    return token, None


def _patch_wls_token(token: str):
    """共用：monkey-patch wls + wf0.get_erp_token 闭包，跟 wf3_logistics_v2 同套路"""
    import sys as _sys
    _hipop_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _hipop_dir not in _sys.path:
        _sys.path.insert(0, _hipop_dir)
    from workflows import wf0_logistics as _wf0
    from workflows import wf_logistics_status as _wls
    _orig = _wf0.get_erp_token
    _wf0.get_erp_token = lambda: token
    _wls.get_erp_token = lambda: token
    return _wf0, _wls, _orig


def _fetch_logistics_nodes(forwarder: str, tracking_no: str) -> dict:
    """playwright 实时抓物流站节点（安时达/阳光/义特/飞坦）。返回 {nodes, note}."""
    if not forwarder or not tracking_no:
        return {"nodes": [], "note": "缺 forwarder 或 tracking_no"}
    try:
        from playwright.sync_api import sync_playwright
        from workflows.wf_logistics_status import query_tracking
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            try:
                r = query_tracking(page, forwarder, tracking_no)
            finally:
                browser.close()
            return r
    except Exception as e:
        return {"nodes": [], "note": f"playwright_error: {type(e).__name__}: {str(e)[:100]}"}


def _physical_tracking_url(forwarder: str, tracking_no: str) -> str:
    """根据 forwarder 拼物流站直链。"""
    if not forwarder or not tracking_no:
        return ""
    from workflows.wf_logistics_status import LOGISTICS_URLS
    base = LOGISTICS_URLS.get(forwarder, "")
    if not base:
        return f"(无直链, forwarder={forwarder})"
    sep = "&" if "?" in base else "?"
    return f"{base}{sep}msg={tracking_no}" if "tracking" in base.lower() else f"{base}{sep}no={tracking_no}"


def _utc_now_iso() -> str:
    import datetime as _dt
    return _dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
























# ── Tool 派发 ─────────────────────────────────────────
# ── 工具实现已外移到 tools_impl（WS-166）。agent.py 仅保留注册/分发/治理入口。
# 重新导出工具实现名，保持 `agent.tool_*` 外部契约（api.py / 测试 / TOOL_FUNCS 投影）不变。
from .tools_impl import (
    tool_query_sku,
    tool_query_order,
    tool_update_alert_status,
    tool_scope_overview,
    tool_compute_replenishment,
    tool_query_replenishment_sku,
    tool_compute_air_freight_roi,
    tool_data_health_check,
    tool_list_products,
    tool_export_table,
    tool_navigate_user_to,
    tool_notify_via_feishu,
    tool_run_workflow,
    tool_query_1688_similar,
    tool_query_sku_live,
    tool_query_order_live,
    _tool_tenant_notes_get,
    _tool_tenant_notes_append,
    _tool_confirm_proposal,
    tool_capture_feedback,
    tool_explain_status_enum,
    tool_query_stock_split,
    tool_total_stock_topn,
)

TOOL_FUNCS = {
    "query_sku": tool_query_sku,
    "query_order": tool_query_order,
    "update_alert_status": tool_update_alert_status,
    "scope_overview": tool_scope_overview,
    "compute_replenishment": tool_compute_replenishment,
    "query_replenishment_sku": tool_query_replenishment_sku,
    "compute_air_freight_roi": tool_compute_air_freight_roi,
    "data_health_check": tool_data_health_check,
    "list_products": tool_list_products,
    "export_table": tool_export_table,
    "navigate_user_to": tool_navigate_user_to,
    "notify_via_feishu": tool_notify_via_feishu,
    "run_workflow": tool_run_workflow,
    "confirm_proposal": lambda proposal_id, user_decision: _tool_confirm_proposal(proposal_id, user_decision),
    "tenant_notes_get": lambda section="": _tool_tenant_notes_get(section),
    "tenant_notes_append": lambda note, section="通用": _tool_tenant_notes_append(note, section),
    "query_sku_live": tool_query_sku_live,
    "query_order_live": tool_query_order_live,
    "query_1688_similar": tool_query_1688_similar,
    "explain_status_enum": tool_explain_status_enum,
    "capture_feedback": tool_capture_feedback,
    "query_stock_split": tool_query_stock_split,
    "total_stock_topn": tool_total_stock_topn,
    }


def _exec_tool(name: str, args: dict, user: dict = None) -> dict:
    """tool 执行前先过 RBAC + governance pipeline。

    - RBAC: user.role → tool 入口权限
    - Governance (Phase 0.2 半 MSCL): destructive tool 走
      ActionProposal → Decision (Haiku) → ExecToken → Execute → ExecutionRecord
      read-only tool 跳过 governance，直调

    ⚠️ INVARIANT (2026-05-26)：
    所有 LLM tool 调用必须经此函数。provider 层（_provider_anthropic /
    _provider_openai）禁止自己实现 _exec_tool —— 历史上 5/21 把 _exec_tool
    复制到 provider 文件只做 RBAC，导致 destructive tool 全部裸跑（governance
    pipeline 形同虚设）。新增 provider 时：from . import agent; agent._exec_tool(...).
    smoke_governance.py 会跑 inspect.getsource 检查 provider 没自定义 _exec_tool。
    """
    try:
        # WS-145 肯定执行意图门:非执行语气（否定/询问/假设/只问影响面）下，
        # LLM 不许偷偷 run_workflow 落任务 —— 在工具执行入口确定性拦死，不靠 prompt。
        if name == "run_workflow":
            _intent = _chat_intent.get()
            if _intent is not None and getattr(_intent, "blocks_llm_execution", False):
                return {
                    "ok": False,
                    "error": "execution_intent_gate_blocked",
                    "blocked_by": "execution_intent_gate",
                    "message": (
                        "本轮是非执行语气（否定/询问/假设/只问影响面），未创建任何后台任务。"
                        "需要执行请明确说「帮我刷新/重算…」。"
                    ),
                }
        from . import rbac as _rbac
        if user and not _rbac.tool_allowed(user, name):
            return {
                "error": "permission_denied",
                "tool": name,
                "user_role": user.get("role"),
                "message": f"当前角色 {user.get('role')} 不能调用 {name}（请向 owner/manager 申请权限）",
            }
        if name not in TOOL_FUNCS:
            return {"error": f"unknown tool: {name}"}
        # ── Governance pipeline（仅 destructive） ──
        from . import governance as _gov
        if _gov.is_destructive(name):
            sc = _chat_scope.get() or {}
            actor = {
                "user_id": (user or {}).get("id") or sc.get("user_id"),
                "email": (user or {}).get("email") or sc.get("current_user_email"),
                "role": (user or {}).get("role") or sc.get("current_role"),
                "tenant_id": (user or {}).get("tenant_id") or sc.get("tenant_id") or _get_tenant(),
                "source": sc.get("source") or "chat",
            }
            return _gov.propose_and_execute(name, args, actor, sc, TOOL_FUNCS)
        # read-only 直调
        fn = TOOL_FUNCS[name]
        return fn(**args)
    except Exception as e:
        return {"error": f"{type(e).__name__}: {e}"}


# ── Chat 主入口 ──────────────────────────────────────
SYSTEM_PROMPT_LEGACY = """你是点购 Agent OS 的店铺协作 Agent，工作在共同空间内（5 个同事 + 1 个你）。

三大原则:
1. 共同工作空间：你和运营/跟单/组长在同一个 chat 里，所有结论和决策可被同事看到、采纳、回滚。
2. 任务泛化：用户可以问任何问题，你应当首选用工具拿到真实数据，再回答。
3. 自主决策：你应该主动给出判断 + 建议（不只是数字），并提供数据出处。

**用户原则**: 用户不应该在终端跑任何脚本。所有数据更新/查询都通过 chat 工具完成。
- 用户问问题 → 你先看现有数据是否够答
- 数据不够新 → 你自动 run_workflow 触发更新（带 followup_prompt 等跑完接续答），不要让用户去跑
- 数据够新 → 直接答 + 给数据出处

当前 scope（参考）:
{scope}

## 问题 → 工具映射（必须遵守）

| 用户问 | 你调的 tool |
|---|---|
| 我该补货吗 / 哪些货要补 / 补多少 / 本周必补 | compute_replenishment |
| <SKU> 卖得怎么样 / 库存够不够 / 趋势 | query_sku |
| **<SKU> 库存拆分 / 四仓 / 义乌仓+沙特仓+noon+在途 / 总库存明细** | **query_stock_split**（必调，不得省略 noon 仓，不得用 TopN 路径）|
| 单 <SKU> 海空运怎么选 / 海运合算还是空运 | compute_air_freight_roi |
| 这单 <PDxxx> 怎么样 / 卡了几天 / 到哪了 | query_order |
| <PDxxx> 已确认丢货 / 已约仓 / 已结案 | update_alert_status |
| 店铺总共多少商品 / 多少 SKU / 多少未上架 | list_products |
| 店铺整体怎么样 / 概览 / 红色告警 | scope_overview |
| 数据是什么时候更新的 / 数据新鲜吗 | data_health_check |
| 跑一下 / 刷新 / 重算 X | run_workflow |
| **导出 / 下载 / 给我 Excel / 给我表格** | **export_table**（必调，不要自己编"已生成 Excel"）|
| **打开 X 页面 / 进 X 模块 / 看 X 看板** | **navigate_user_to**（必调，禁编虚构 URL）|
| **发飞书 / 通知刘鹤 / 推到群里 / @同事** | **notify_via_feishu**（必调，禁说"已发到飞书"）|

## 数据新鲜度自动判断（**所有问题**都遵守这个流程）

**核心**：每个用户问题都有上游依赖。回答之前先确认**所有依赖源**新鲜，不能只看终端表（如 wf5）。

### ⚠️ 强制规则（避免 hallucinate — 这些是会被运营当面骂"骗子"的事故）

**1. 任何业务数据回答前，必须先调 `data_health_check` 拿真实新鲜度**
   - 不要凭空猜"noon 销量是 X 天前"——必须从 tool 返回读
   - 不要假设字段值——所有数字/日期/SKU id 都要来自 tool 返回
   - 不要抄本 SYSTEM_PROMPT 的举例数字（举例里的"5 月 4 日"、"3 天前"全是占位符示意）

**2. 严禁宣称做了"未真正做"的事**
   - "✅ 已触发导出/同步/刷新/通知" → 必须真有对应 tool 调用并返回成功才能这么说
   - "已为你生成 Excel/链接" → **本系统没有 Excel 导出功能**，禁止编造
   - "已为你打开 X 页面" → 你不能打开页面，只能告诉用户"在工作台 sidebar 找 X 模块"
   - "我刚调用了 X 工具" → 当且仅当本轮真的调用了，否则不要写

**3. 严禁编造 URL / 域名 / 页面元素 / UI 按钮 / UI 操作路径**
   - 不要编 `https://agent.diangou.ai/...` 这种**虚构域名**
   - 不要描述前端不存在的 UI（"顶部 Tab 高亮"、"右上角导出按钮已激活"、"行末 🔍 按钮"）
   - **严禁编"在 sidebar/侧边栏 找到 X 按钮 → 点 Y"** —— sidebar 真实菜单只有：今日总览 / 数据获取 / 销售-库存 / 在途物流 / 补货决策 / 流量推广 / 选品+货源 / 营销活动 / 飞书沉淀 / 数据巡检 / 跟单跨店 + 系统块（Agent 操作记录 / 策略沉淀 / 数据刷新）。**绝不要描述"侧边栏的某某子菜单/路径/选项"，因为模型对实际 DOM 的猜测 80% 是错的**
   - 工作台真实的模块只有：overview / sales / logistics / replenish / selection / feishu / audit + role/liuhe，路径都是 localhost:8765/module/<name>
   - 真有的入口才能引导用户去；不确定就让用户"sidebar 看一下"

**3b. 用户问"刷新 / 跑工作流 / 同步数据 / 重算 X / 扫 ERP / 拉数据"时，必须本轮**真的**调 `run_workflow`，禁止只口头描述**
   - 你**有** run_workflow tool，能直接触发后台跑（前端会自动订 SSE 显进度）
   - "扫 / 拉 / 同步 / 刷新 / 重算 / 跑一下" 都是同一类动词 —— 必须 run_workflow，不能假装"再次触发"
   - **死规矩**：本轮你说出"已触发 / 已启动 / 已开始 / 再次触发 / 系统已经在后台跑了" 等表述 ⇔ 本轮 tool_use 块里必须有 `run_workflow` 实际调用。两者必须同时为真；只说不调 = 撒谎 = 事故
   - 用户连发两次同一指令时，**不要**假设"上次已触发了"（你不知道上次有没有真触发）—— 重新调 run_workflow 一次，最多重复了一次，比让用户以为任务在跑实际没跑要好
   - 禁说"这个需要组长/管理员账号才能触发" / "我没有权限" / "Agent 当前没有权限" —— 你已经被赋予 run_workflow，能跑就跑；只有 tool 真返回 `permission_denied` 才能这么回
   - 禁说"在工作台 sidebar 找到 X → 点 Y" —— 这种 UI 路径几乎必编错；直接 run_workflow 就对了

**3c. destructive tool 返回 action_type='plan' 时 — Explore→Plan→Implement 三段**
   - 高风险 destructive（update_alert_status 改物流告警 / 等）走治理 pipeline，第一次调返
     一个 dict 含 action_type='plan' + plan_text + proposal_id 字段
   - 你必须**原文转告 plan_text 给用户**，让用户回 OK / 不要 / 改
   - 用户回 "OK / 是 / 确认" → 本轮必须调 `confirm_proposal(proposal_id=..., user_decision='ok')`
   - 用户回 "不要 / 取消 / no" → 调 `confirm_proposal(proposal_id=..., user_decision='cancel')`
   - **绝不要**自己再次调原 destructive tool（governance 会拒）

**4. 用户报告状态变化时（如"我刷新了"、"我上传了"），必须重新调 tool 验证**
   - 不要直接信用户的报告就回"已确认更新"
   - 调 data_health_check / get_data_health 看真实 stale_days
   - 如果用户说更新了但实际没更新，要明确告诉用户"我看到的还是 X 天前，可能你的上传还没 ingest 完，或者文件没识别出来"

**5. 时间戳精度禁忌**
   - data_health_check 返回的日期都是 `YYYY-MM-DD` 粒度，**没有时分秒**
   - 严禁编造 `14:22:07Z` / UTC 偏移 / 沙特时间换算这种伪精确时间戳
   - 如果工具返回 `2026-05-05`，你只能说"5 月 5 日"，不能扩展成"2026-05-05T14:22:07Z UTC（沙特时间 17:22）"

**6. 表格字段必须用真实存在的列**
   - 现有 wf5 字段：`partner_sku / trend / daily_rate / urgency / weekly_total_replenish / current_pipeline / target_pipeline / ops_advice / risk_label / sellable_days / decision_days`
   - 现有 wf2 字段：`partner_sku / title / sales_10d / sales_30d / sales_60d / sales_90d / sales_180d / latest_price / latest_profit_rate / is_listed / sales_grade`；query_sku 工具额外返回 `total_orders_30d`（30d 窗口总单）/ `cancel_rate_30d`（30d 取消率）/ `return_rate_30d`（30d 退货率）/ `history_total`（ERP 历史总单）/ `as_of_date`（数据口径截止日）；快照超期时额外返回 `data_stale=True`（快照超过 3 天）/ `stale_days`（快照距今天数，仅在 data_stale=True 时出现）
   - **快照时效规则**：工具返回 `data_stale=True` 时（快照超过 3 天，或 `as_of_date` 为空）：① 必须明确告知用户数据已过期（"数据已超过 X 天，可能不是最新"）；② 不得把过期快照里的销量/取消率/退货率当作当前事实直接报出；③ 建议用户刷新（触发 run_workflow 重新 ingest，或上传最新 noon CSV）。`data_stale` 字段不存在时（3 天内）：直接用快照数据回答，附带 `as_of_date` 供用户参考
   - **严禁这些不存在的中文字段名**（已是反复事故源）：
     - ❌ "可撑天数" → 用 `sellable_days`（数据库真名）
     - ❌ "7 天销量" → 用 `sales_10d`（最近的真实窗口；没有 7 天）
     - ❌ "海运 ROI 预估" / "空运 ROI 预估" / "推荐物流方式" → 这些只能在调用了 `compute_air_freight_roi` 工具后才能引用
     - ❌ "可售周期" / "周转天数" / 任何 wf5 字段表里没有的中文名 → 不要用
   - 想要的字段如果工具返回里没有，直接说"这个字段我们目前不算"而不是编一个数

### 流程

1. **识别意图**（intent）：把用户问题映射到一种 intent
2. **拿依赖源**：调 `data_health_check` → `dependency_groups[intent]` → 列出该意图依赖的所有源
3. **检查每个源**：用 tool 返回的 `sources[<source>].stale_days` 和 `automation`
4. **行动**：
   - 全新鲜（< stale_threshold_days）→ 调对应查询 tool 直接答
   - **automation=auto 陈旧** → run_workflow(对应 workflow) + followup_prompt（用户原始问题）
   - **automation=needs_csv 陈旧** → **不要** run_workflow，给精确上传指引（path + csv_pattern），引导用户上传到工作台 📤 区
   - 混合 → 先列上传引导（needs_csv 部分），auto 部分一并 run_workflow

### 意图 → 依赖源 + 推荐 tool（必背）

| 用户说 | intent | 依赖源 | 数据齐了调什么 tool |
|---|---|---|---|
| 我该补货吗 / 哪些要补 / 补多少 / 本周必补 | `replenishment` | erp_sales + erp_stock + noon_orders + noon_stock + wf3_logistics + wf5_replenish | compute_replenishment |
| `<SKU>` 卖得怎么样 / 趋势 / 库存够不够 | `sku_health` | erp_sales + noon_orders + wf3_logistics + wf5_replenish | query_sku |
| 在途 / 物流追踪 / 货到哪了 | `logistics_track` | wf3_logistics | query_order 或 scope_overview |
| 告警 / 卡单 / 红色货单 / `<PDxxx>` | `alerts` | wf3_logistics + wf6_alerts | query_order |
| 单 SKU 海运空运怎么选 | `air_freight_roi` | erp_sales + noon_orders + wf5_replenish | compute_air_freight_roi |
| 店铺总共多少商品 / SKU 数 / 未上架 | `products_count` | erp_products | list_products |
| 店铺整体怎么样 / 概览 | `overview` | erp_sales + wf3_logistics + wf5_replenish + wf6_alerts | scope_overview |
| 销量 X 天卖了多少 | `sales_only` | erp_sales + noon_orders | query_sku 或 list_products |
| 库存够不够 / 还能撑几天 | `stock` | erp_stock + noon_stock | query_sku |
| 数据新鲜吗 / 什么时候更新的 | （直接答） | — | data_health_check |
| 跑/刷新/重算 X | （直接触发） | — | run_workflow |
| `<PDxxx>` 已确认丢货 / 已结案 | （写入） | — | update_alert_status（要确认意图） |

### 上传引导话术（needs_csv 陈旧时）

不要泛泛说"去上传 CSV"，要给精确指引（来自工具返回的 `sources[<src>].where` + `csv_pattern`）。

**模板（具体数字必须从工具返回值里取，不要凭空编造日期或数字）**：

> 你 [STORE] 的 [源中文名] 是 [N] 天前的（最新到 [日期]），我不能自动刷新这部分。
>
> 👉 请操作：
> 1. [where 字段的导出路径]，文件名形如 `[csv_pattern]`
> 2. 拖到工作台**顶部 📤 上传区**
>
> 上传完会自动 ingest + 重算，跑完我会接着告诉你『[用户原始问题]』的最终答案。

### 多个 needs_csv 源都陈旧时

合并指引（一次告诉用户全部要传的 CSV），不要分多次。

### 混合陈旧（auto + needs_csv）

例：用户问补货，noon_orders 陈旧 + erp_stock 陈旧。
- 告诉用户上传 noon CSV（needs_csv 部分要人工）
- 提一句"ERP 库存我会在你上传后顺便刷新"
- **不要先 run_workflow(wf1_stock)**，因为最终 wf5 还要等 noon 数据来才能正确算，单独跑 wf1 是浪费

### 已经触发后

run_workflow 调完后**不要**再 query 数据。前端会在跑完自动重发用户原始问题（followup_prompt），那时再用最新数据答。

### 用户坚持用旧数据时（关键场景）

用户可能不想等更新，要立刻拿当下数据。识别信号：
- 直接说："就用现在的" / "不用更新" / "先看看" / "凑合给个" / "粗略估" / "我现在就要"
- 拒绝上传 / 拒绝跑 workflow：在你给完上传指引或触发建议后，用户重复问同样问题或说"先告诉我"
- 上下文暗示赶时间："5 分钟后开会，告诉我"

**这种情况你应该**:
1. **不要**坚持要求更新，直接用旧数据答
2. **必须明确警示**：在答案开头一句话告诉用户具体哪些源陈旧多少天，结论可能因此偏向哪个方向（如"noon 销量数据是 4 天前的，最近一周的爆款会被低估，结论偏保守"）
3. 调对应查询 tool（compute_replenishment / query_sku / 等），照常给数据 + 出处
4. **结尾**附一句"如要更准的结论，跟我说『刷新数据』或上传最新 CSV"

**陈旧偏向参考**（用来给警示）:
- noon_orders 陈旧 → 漏掉最近订单 → 销量低估、利润率以历史为准、新爆款看不到
- noon_stock 陈旧 → 平台库存可能更紧张/更宽松 → 库存可撑天数有偏差
- erp_stock 陈旧 → 国内仓和海外仓库存数有偏差
- wf3_logistics 陈旧 → 在途到货时间预估不准
- wf5_replenish 陈旧 → 补货建议是上次跑的快照（如果 wf2/wf1/wf3 都新但 wf5 旧，可以 run_workflow(wf5_sales_cycle) 快速重算，不需要等）

**例子**：
- 用户："不用上传 CSV 了，就用现在的告诉我哪些要补"
- 你："好的。⚠️ noon 销量是 4 天前数据，结论偏保守（最近一周的爆款会被低估）。
       基于现有数据：本周必补 X 个 SKU... [给数据] 📎
       如需更准结论，上传最新 sales_noon_*_KSA_*.csv 后重问。"

## 回答风格

- 中文，简洁，2-4 句一段，不要罗列冗长字段
- 给判断（趋势 / 紧迫度）+ 简明建议（量化、可执行）
- 不知道时直说，不要瞎编
- 涉及写入（update_alert_status）需要用户确认意图后再调用
- run_workflow 不需要二次确认（页面有进度条），直接调
- 触发 run_workflow 后**不要再 query 数据**，等 followup_prompt 自动接续

## 输出风格 — 思考过程的"业务化简化"

**鼓励一句话的业务进度提示**（让用户感知 Agent 在做事）：
- ✅ "我先看看店铺整体情况"
- ✅ "我来查一下补货建议"
- ✅ "稍等，我对一下数据"

**绝不要暴露技术细节/内部字段名**：
- ❌ "这个问题属于 replenishment 意图"
- ❌ "依赖源 noon_orders.stale_days=3，automation=needs_csv"
- ❌ "我调用 data_health_check / compute_replenishment tool"
- ❌ "首先 X，然后 Y，接下来 Z" 的多步骤罗列

**陈旧警示也用业务语言**：
- ✅ "noon 销量是 4 天前，结论偏保守"
- ❌ "noon_orders source stale_days=4，automation=needs_csv"

## 例子对照

❌ 错（技术细节满天飞）：
> 这个问题属于 replenishment 意图，依赖 6 个源。
> 我先调 data_health_check 检查 stale_days...
> 看到 noon_orders.automation=needs_csv，stale_days=3，需要上传。

✅ 对（一句业务进度 + 直接结论）：
> 我来看看你的补货情况。
> noon 销量是 3 天前的，我不能自动刷新这部分。
> 👉 请到紫鸟 noon 后台 sales 页面 export 最近 180 天 CSV，拖到工作台 📤 上传区。
> 上传后我会接着告诉你哪些要补。
"""


# Phase 0.3 Context Engineering — SYSTEM_PROMPT 砍到 ~1500 token（4188→1500，节省 65%）
# 按 Anthropic Claude Code Best Practices："ruthlessly prune"、"too long → ignored"
# 多数细则已经被结构性约束做了（_safety hook / governance pipeline / smoke），
# prompt 只留 minimal essentials.
SYSTEM_PROMPT = """你是点购 Agent OS 的店铺协作 Agent。

scope: {scope}

## 工作流
1. 业务问题先调 data_health_check 看新鲜度 → 再调对应查询 tool 答
2. 数据陈旧 → run_workflow（auto 类）/ 给上传指引（noon CSV 类）
3. destructive tool 返回 action_type='plan' → 原文转告 plan_text 让用户回 OK → 调 confirm_proposal(pid,'ok')

## 关键：用户问"某 SKU/某货单当前在途 / 物流状态"时，**直接调 query_sku_live / query_order_live**
- 不要先 data_health_check 然后说"wf3 陈旧，等 ingest 完再答"——这是错的，应该跳过 wf3 缓存直接查 ERP
- 不要说"我可以查"然后不调 tool —— 必须本轮真调 query_sku_live(sku=...)
- 用户问多个 SKU → 对每个分别调 query_sku_live
- query_sku_live 慢（5-15s）但准（直连 ERP），值得
- query_sku_live 返回 `ok=false`（例如 `erp_login_failed_no_cache`，且 `cache_fallback=false`）时：必须明告用户「ERP 实时不可用，无法确认当前在途」，**不许**把 wf3 旧缓存当作实时数据呈现

## tool 速查
| 用户问 | 调 |
|---|---|
| <SKU> 卖得 / 库存 / 趋势 | query_sku |
| <SKU> 当前在途多少 / 实时物流 / wf3 陈旧时 | **query_sku_live**（实时 ERP，5-15s）|
| <PDxxx> 状态 / 卡几天（hipop 缓存）| query_order |
| <PDxxx> 现在到哪了 / 实时状态 / 物流码 | **query_order_live**（实时 ERP）|
| <PDxxx> 已确认丢货/已结案 | update_alert_status |
| 我该补货吗 / 必补哪些 | compute_replenishment |
| 海运空运 ROI | compute_air_freight_roi |
| 店铺概览 / 红色告警 | scope_overview |
| 商品总数 / SKU 数 | list_products |
| 数据新鲜吗 | data_health_check |
| 跑/刷新/扫/拉/重算/同步 | run_workflow |
| 导出/下载/Excel/打成表格 | **export_table**（真生成 xlsx，filtered_count 才是真总数；返完用 [文件名](download_url) markdown 给用户）|
| 状态/字段哪来 / 5 个状态出处 / 能加 X 状态吗 / 是 ERP 字段还是 hipop 字段 | **explain_status_enum**（不要凭空说"系统写死"，必须真调拿 source 引用）|
| 打开 X 页面 | navigate_user_to |
| 总库存最高 / 库存最多 / 积压最多 / 库存 TopN | **total_stock_topn**（含 pending_inbound，口径 = noon+海外+国内+送仓未上架；超 3 天 fail-closed；与 noon 可售 saleable 不同）|
| 撞到你做不了/超范围 → 用户回"记一下/提个需求/帮我记" | **capture_feedback** |

## 撞限即捕获需求（WS-26）
- 用户确认（记一下/好/提个需求）→ 本轮必须调 capture_feedback(content=用户诉求原话)。
- capture_feedback 返 ok=False → 如实说"没记成，等会儿再说一次"，**绝不**假装记了。

## 死规矩（违反 = 事故）
1. **业务数据先调 data_health_check**，不要凭空猜"X 天前更新"
2. **SKU id / 数字必须来自 tool 返回**；工具未返回的值直接说"目前不算"
3. **用户报告状态变化（"我刷新了"/"我传了"）必须重新调 tool 验证**，不信用户报告
4. **data_stale=True 时禁止报数值**：query_sku 返回 `data_stale=True`（快照超 3 天或无 as_of_date，此字段仅在过期时出现），所有数值字段已 REDACT 为 null；你必须如实告知数据过期，禁止从工具返回或上下文中"估算"已 REDACT 的旧值，也不得附免责声明后仍报旧值

## 长期偏好沉淀
- 用户明确说"以后都这么办" / "记住" / "默认 X" → 调 tenant_notes_append
- 高风险决策前可调 tenant_notes_get 看客户既定偏好（按需，不每次都拉）

## 回答风格
- 中文 2-4 句一段，给结论 + 简明建议
- 一句进度 OK（"我先看看"），不暴露技术细节（"调用 X tool"）
- run_workflow 后不再 query，等 followup_prompt 自动续

## 数据陈旧场景
- 用户说"就用现在的 / 不用更新" → 直接答 + 开头一句警示（"noon 销量是 X 天前，结论偏保守"）+ 末尾一句"如要更准，跟我说刷新"
- noon CSV 陈旧 → 给精确路径 + 拖工作台 📤 区（不是 run_workflow）
- ERP 陈旧 → run_workflow + followup_prompt 自动续答
"""


_JUDGE_SYSTEM_PROMPT = (
    "你是 Agent 回复质量评判官。基于用户问题、Agent 回复、调用的工具、系统检测到的幻觉信号，"
    "判断这个回复是否真回答了问题、有无编造、引用是否支撑结论。\n"
    "严格只返回 JSON（不要任何其他文字）：{\"confidence\": 0~1 浮点, \"verdict\": \"一句话评判\"}。\n"
    "工具调用多、有数据引用、无幻觉信号 → 高置信(0.8+)；凭空作答、有幻觉信号 → 低置信(0.4-)。"
)


def _run_llm_judge(question, reply, tool_log, warnings):
    """独立 LLM 给回复打分。复用 governance 的 LLM 调用 + JSON 抽取 pattern。
    走当前 provider（默认 deepseek，便宜）。失败返 None → 调用方退回启发式分。"""
    from . import _provider, governance as _gov
    prompt = (
        f"用户问：{(question or '')[:300]}\n\n"
        f"Agent 回复：{(reply or '')[:600]}\n\n"
        f"调用工具：{[t['name'] for t in tool_log]}\n"
        f"系统检测到的幻觉信号：{warnings or '无'}\n\n"
        "严格返回 JSON。"
    )
    try:
        r = _provider.chat_with_tools(
            messages=[{"role": "user", "content": prompt}],
            system=_JUDGE_SYSTEM_PROMPT, tools=[], tool_funcs={},
            scope={"_judge_only": True},
        )
        return _gov._extract_json(r.get("reply") or "")
    except Exception:
        return None


def _compute_judge_confidence(question, reply, tool_log, refs, warnings):
    """judge + confidence 混合算法。
    启发式 baseline 每次跑（0 成本）；低置信 OR destructive 时触发 LLM judge 复核。
    返回 (judge_text, confidence_float, method)。
    """
    from . import governance as _gov
    n_tools = len(tool_log)
    n_fields = sum(len(t.get("result_keys") or []) for t in tool_log)
    n_warn = len(warnings or [])
    has_refs = bool(refs)

    # 启发式打分
    conf = 0.85
    conf -= 0.15 * min(n_warn, 3)        # 幻觉信号惩罚（最多 -0.45）
    if n_tools == 0: conf -= 0.20        # 凭空作答（没调任何 tool）
    if not has_refs: conf -= 0.10        # 无数据源引用
    conf = max(0.1, min(conf, 0.95))

    parts = [f"{n_tools}工具/{n_fields}字段"]
    if n_warn: parts.append(f"{n_warn}个幻觉信号")
    if refs: parts.append("源:" + ",".join((r.get("table") or "")[:20] for r in refs[:2]))
    judge = " · ".join(parts)[:200]
    method = "heuristic"

    # 混合：低置信 OR destructive tool → LLM judge 复核
    is_destr = any(_gov.is_destructive(t["name"]) for t in tool_log)
    if conf < 0.6 or is_destr:
        llm = _run_llm_judge(question, reply, tool_log, warnings)
        if llm and "confidence" in llm:
            try:
                conf = max(0.1, min(float(llm["confidence"]), 0.99))
                judge = (llm.get("verdict") or judge)[:200]
                method = "llm"
            except (TypeError, ValueError):
                pass  # LLM 返回的 confidence 不是数字 → 保留启发式分
    return judge, conf, method


import re as _re
# _safety 加的 banner / 低置信 tip 是给用户的展示层，绝不能回流进 LLM 历史 —— 否则
# LLM 复读 banner 文字（含"之前触发任务还在跑"触发词）→ sanitize 再包一层 → 无限自激双 banner。
_SAFETY_BANNER_RE = _re.compile(
    r"⚠️ \*\*系统检测到 Agent 回复中可能存在不准确之处\*\*：[\s\S]*?"
    r"以下是原始回复（已标记可疑部分）：\s*---\s*"
)
_LOWCONF_TIP_RE = _re.compile(r"⚠️ 我对这个回答的置信度较低（\d+%）[^\n]*\n\n---\n\n")
_INTENT_EXPLAIN_RE = _re.compile(r"\*{0,2}本轮我先不动手\*{0,2}[（(]你是在问能不能[）)]|[（(]\*{0,2}本轮不执行\*{0,2}[）)]|按你说的\*{0,2}不执行\*{0,2}这步刷新|本轮未执行。需要执行请明确说")

def _strip_safety_banner(text):
    """剥掉 _safety banner + 低置信 tip，拿回干净正文（用于持久化 + 喂 LLM 历史）。"""
    if not text or not isinstance(text, str):
        return text
    text = _SAFETY_BANNER_RE.sub("", text)
    text = _LOWCONF_TIP_RE.sub("", text)
    return text


def _clean_history(messages: List[Dict]) -> List[Dict]:
    """喂 LLM 前清掉 assistant 历史里残留的 banner（清 DB 已有的脏数据 + 防自激）。"""
    out = []
    for m in messages:
        if m.get("role") == "assistant" and isinstance(m.get("content"), str):
            c = _strip_safety_banner(m["content"])
            out.append({**m, "content": "[上轮：询问/假设/影响面，未执行操作]" if _INTENT_EXPLAIN_RE.search(c) else c})
        else:
            out.append(m)
    return out


# ── 反馈/需求捕获：撞限时确定性补一句 offer（WS-26）──────────────
# 验收①要求 agent 回复『做不到/超范围』时**必含**一句 offer。靠 prompt 提醒不可靠，
# 这里做成确定性后处理 hook：只认 Agent 自述能力受限的措辞（第一人称做不了 / 功能不支持），
# 不碰数据陈述（如『库存撑不到下周』），避免污染正常回答路径（验收④）。
_DEADEND_RE = _re.compile("|".join([
    r"我(暂时|目前|这边|现在)?(做不了|做不到|无法|没法|帮不了|处理不了|实现不了|搞不定)",
    r"(帮不了|做不了|做不到|满足不了|无法满足|无能为力)(你|您)",
    r"超出(我|我的|当前|目前|系统)?.{0,4}(能力|范围|权限)",
    r"不在(我|我的|当前|目前)?.{0,6}(能力|范围|功能)(范围|内|之内)?",
    r"(这个|该|此)?功能(暂时|目前)?(还)?(不支持|没有|未上线|做不了|做不到)",
    r"系统(暂时|目前)?(还)?(不支持|没有这个功能|不具备|做不了)",
    r"暂(时)?(不|未)支持",
    r"目前(还)?(做不到|不支持|无法|没有这个)",
]))
# offer 标记串：_OFFER_MARK 是 hook 补的话术里的固定串；_OFFER_SEEN 用于判定 LLM
# 是否已经自己 offer 过（避免重复补 —— LLM 常自带「要我记成需求吗」）。
_OFFER_MARK = "记成一条需求"
_OFFER_LINE = "💡 要我把它记成一条需求反馈给产品吗？你回「记一下」我就转过去。"
_OFFER_SEEN = ("记成需求", "记成一条需求", "记成反馈", "记成一条反馈",
               "提个需求", "记下来当需求", "记录成需求")


def _needs_feedback_offer(reply: str, tools_used: List[str]) -> bool:
    """reply 表达了『我做不了/超范围』，且本轮没调过 capture_feedback、reply 里也还没 offer。"""
    if not reply or not isinstance(reply, str):
        return False
    if "capture_feedback" in (tools_used or []):
        return False            # 已经在记了，别再 offer
    if any(m in reply for m in _OFFER_SEEN):
        return False            # LLM 自己已经 offer 了，不重复
    return bool(_DEADEND_RE.search(reply))


def _maybe_append_feedback_offer(reply: str, tools_used: List[str]) -> str:
    """撞限回复确定性补一句 offer；正常回复原样返回。"""
    if _needs_feedback_offer(reply, tools_used):
        return reply.rstrip() + "\n\n" + _OFFER_LINE
    return reply


def _maybe_append_stock_readiness_warning(reply: str) -> str:
    status = _last_replenishment_stock_status.get()
    if not status or status.get("ready"):
        return reply
    text = reply or ""
    if any(k in text for k in ("未更新", "不新鲜", "滞后", "偏保守", "旧数据", "数据旧", "库存旧")):
        return reply
    warning = "提示：库存数据未更新或不完整，当前补货结论偏保守；请先完成库存更新后再计算。"
    return text.rstrip() + "\n\n" + warning


def _ensure_export_download_link(reply: str, tool_log: list) -> str:
    """If export_table produced a file, make the real /api/download link visible."""
    text = reply or ""
    for t in (tool_log or []):
        if t.get("name") != "export_table":
            continue
        url = t.get("result_download_url")
        if not url or url in text:
            continue
        filename = t.get("result_filename") or url.rsplit("/", 1)[-1] or "导出文件.xlsx"
        if "(download_url)" in text:
            text = text.replace("(download_url)", f"({url})")
        else:
            text = text.rstrip() + f"\n\n下载链接：[{filename}]({url})"
    return text


def _maybe_inject_missing_rates(reply: str, question: str) -> str:
    """
    确定性后注入：用户问取消率/退货率时，若 LLM 回复未包含该数值，
    从 _last_sku_rate_stats 提取 pct 字段并追加，防止遗漏。
    数据过期（data_stale=True）或字段为 null 时追加说明缺失原因。
    """
    q = (question or "").lower()
    wants_cancel = any(x in q for x in ("取消率", "cancel_rate"))
    wants_return = any(x in q for x in ("退货率", "return_rate"))
    if not (wants_cancel or wants_return):
        return reply
    items = _last_sku_rate_stats.get()
    if not items:
        return reply
    text = reply or ""
    injected = []
    for item in items:
        if not item.get("found"):
            continue
        sku = item.get("sku", "")
        data_stale = item.get("data_stale", False)
        if data_stale:
            if wants_cancel and "取消率" not in text and "cancel" not in text.lower():
                stale_days = item.get("stale_days", 0)
                injected.append(f"{sku} 取消率数据已过期（{stale_days} 天未刷新），无法给出准确数值")
            continue
        cancel_pct = item.get("cancel_rate_30d_pct")
        return_pct = item.get("return_rate_30d_pct")
        if wants_cancel and cancel_pct:
            num_str = cancel_pct.rstrip("%")
            if num_str not in text:
                injected.append(f"{sku} 取消率（30d）：{cancel_pct}")
        elif wants_cancel and cancel_pct is None and "取消率" not in text:
            injected.append(f"{sku} 取消率数据暂缺，请确认 wf2_orders 是否已导入")
        if wants_return and return_pct:
            num_str = return_pct.rstrip("%")
            if num_str not in text:
                injected.append(f"{sku} 退货率（30d）：{return_pct}")
    if not injected:
        return reply
    return text.rstrip() + "\n\n（补充）" + "；".join(injected)


def _asks_workflow_impact(question: str) -> bool:
    q = question or ""
    return any(
        marker in q
        for marker in (
            "会更新哪些表",
            "更新哪些表",
            "会更新哪些数据",
            "更新哪些数据",
            "影响哪些数据",
            "影响哪些",
            "影响面",
        )
    )


def _workflow_business_impact_reply(workflow: str, question: str) -> str:
    if workflow != "wf1_stock_v2" or not _asks_workflow_impact(question):
        return ""
    return (
        "影响面：会刷新 ERP 6 仓库存快照；补货建议、售罄天数和补货判断"
        "会使用新库存重算。"
    )


_DATA_HEALTH_DATE_RE = _re.compile(r"5\s*月|2026-05|\b05-\d{2}")
_DATA_HEALTH_QUESTION_RE = _re.compile(r"数据.{0,12}(?:什么时候|何时|更新|新鲜)|(?:什么时候|何时).{0,12}数据|更新的数据")


def _maybe_append_oldest_data_health_date(
    reply: str, question: str, tools_used: List[str], scope: Dict
) -> str:
    if "data_health_check" not in tools_used or not _DATA_HEALTH_QUESTION_RE.search(question or ""):
        return reply
    if _DATA_HEALTH_DATE_RE.search(reply or ""):
        return reply
    try:
        health = _data.get_data_health((scope or {}).get("store", "KSA"))
    except Exception:
        return reply

    labels = {
        "erp_products": "ERP 商品",
        "erp_sales": "ERP 销量",
        "erp_stock": "ERP 库存",
        "noon_orders": "noon 销量订单",
        "noon_stock": "noon 库存",
        "wf3_logistics": "物流数据",
        "wf5_replenish": "销售周期/补货决策",
        "wf6_alerts": "物流告警",
    }
    dated = []
    for key, source in (health.get("sources") or {}).items():
        latest = str((source or {}).get("latest") or "")[:10]
        if _re.match(r"\d{4}-\d{2}-\d{2}$", latest):
            dated.append((latest, labels.get(key, key)))
    if not dated:
        return reply
    oldest_date, oldest_label = min(dated, key=lambda item: item[0])
    return (
        (reply or "").rstrip()
        + f"\n\n补充：当前最旧的数据来源是{oldest_label}，最新日期 {oldest_date}。"
    )


_ORDER_NEGATIVE_HINT_RE = _re.compile(r"未找到|不存在|无物流|无记录|找不到|核实货单号")
_ORDER_BLOCKER_SHAPED_RE = _re.compile(
    r"没有.{0,10}(?:单号|货单号)|不像.{0,12}货单号|无法.{0,12}(?:查询|复核)|ERP.{0,20}(?:账号|凭据).{0,10}(?:未|没|无)"
)


def _maybe_append_order_lookup_negative_hint(reply: str, question: str, tools_used: List[str]) -> str:
    if "query_order_live" not in tools_used:
        return reply
    if _ORDER_NEGATIVE_HINT_RE.search(reply or ""):
        return reply
    if not _ORDER_BLOCKER_SHAPED_RE.search(reply or ""):
        return reply
    return (
        (reply or "").rstrip()
        + "\n\n补充：请核实货单号；当前没有可用 ERP 实时记录。"
    )


def _maybe_append_navigation_url(reply: str, tool_log: List[Dict]) -> str:
    if "localhost:8765" in (reply or ""):
        return reply
    for tool in tool_log or []:
        if tool.get("name") != "navigate_user_to":
            continue
        raw_args = tool.get("args") or {}
        if isinstance(raw_args, str):
            try:
                args = json.loads(raw_args)
            except Exception:
                args = {}
        elif isinstance(raw_args, dict):
            args = raw_args
        else:
            args = {}
        module = args.get("module")
        if not module:
            return reply
        nav = tool_navigate_user_to(module, args.get("store") or "KSA")
        if not nav.get("ok"):
            return reply
        return (reply or "").rstrip() + f"\n\n入口：{nav['url']}"
    return reply


from ._deterministic_routes import (  # WS-167: 确定性路由/formatter 外移到非锁模块
    _deterministic_data_freshness_request,
    _deterministic_erp_refresh_time_request,
    _deterministic_export_request,
    _deterministic_multi_workflow_request,
    _deterministic_product_sales_topn_request,
    _deterministic_products_count_request,
    _deterministic_readonly_reply,
    _deterministic_readonly_request,
    _deterministic_replenishment_list_request,
    _deterministic_replenishment_sku_request,
    _deterministic_scope_overview_request,
    _deterministic_sku_metric_request,
    _deterministic_stock_split_request,
    _deterministic_total_stock_topn_request,
    _deterministic_workflow_request,
    _extract_live_order_no,
    _fmt_int,
    _format_data_freshness_reply,
    _format_erp_refresh_time_reply,
    _format_metric_value,
    _format_order_live_reply,
    _format_pct,
    _format_product_sales_topn_reply,
    _format_products_count_reply,
    _format_replenishment_list_reply,
    _format_scope_overview_reply,
    _format_sku_metric_reply,
    _format_stock_split_reply,
    _format_total_stock_topn_reply,
    _has_stock_refresh_intent,
    _stock_refresh_refusal_reply,
    _stock_refresh_refused,
)


def _current_workflow_task(workflow: str) -> Optional[dict]:
    try:
        rows = _data._fetch(
            "SELECT task_id, workflow, state FROM tasks WHERE tenant_id=? AND workflow=? "
            "AND state IN ('running','queued') ORDER BY COALESCE(last_heartbeat, started_at) DESC LIMIT 1",
            (_get_tenant(), workflow),
        )
        return dict(rows[0]) if rows else None
    except Exception:
        return None


_RUNNING_WORKFLOW_TASK_RE = re.compile(
    r"已有运行中实例:\s*\[\s*['\"]?([0-9a-fA-F]{8})"
)


def _existing_workflow_task_id(tool_result: dict) -> Optional[str]:
    """Return the real task id when governance denies only because the workflow is already running."""
    if not isinstance(tool_result, dict) or tool_result.get("action_type") != "denied":
        return None
    reason = tool_result.get("reason") or ""
    m = _RUNNING_WORKFLOW_TASK_RE.search(reason)
    return m.group(1).lower() if m else None


def _workflow_registry_summary(workflow: str, fallback_label: str) -> tuple[str, int, list]:
    try:
        from . import api as _api
        label, steps, affected = _api.WORKFLOW_REGISTRY.get(workflow, (fallback_label, [], []))
        return label, len(steps), affected
    except Exception:
        return fallback_label, 0, []


def _active_workflow_task(workflow: str) -> Optional[Dict]:
    rows = _data._fetch(
        "SELECT task_id, workflow, state FROM tasks "
        "WHERE tenant_id=? AND workflow=? AND state IN ('running', 'queued') "
        "ORDER BY COALESCE(last_heartbeat, started_at) DESC NULLS LAST LIMIT 1",
        (_get_tenant(), workflow),
    )
    if not rows:
        return None
    from . import api as _api
    label, steps, affected = _api.WORKFLOW_REGISTRY.get(workflow, (workflow, [], []))
    task = rows[0]
    return {
        "task_id": task["task_id"],
        "workflow": workflow,
        "label": label,
        "total_steps": len(steps),
        "affected_modules": affected,
        "followup_prompt": None,
        "state": task.get("state"),
    }


def _logistics_task_evidence_check(task_id: str) -> Optional[str]:
    """物流入口证据检查（T21-SUB-3）：用 SUB-1 统一回读接口验证 durable 任务证据。

    返回 None → 证据完整（task row 存在 + ≥1 queued/started 事件），无需降级。
    返回字符串 → 证据缺失，调用方将此字符串作为降级回复（替换「已触发」）。

    任务表报错或事件缺失 → 回复降级为「未确认创建成功」，绝不返回假成功。
    孤儿事件（agent_events 有记录但 tasks 行不存在）同样降级，不放行假成功。
    """
    try:
        evidence = _data.get_task_with_events(task_id)
    except Exception:
        return ("物流后台任务**未确认创建成功**（任务表查询出错）。"
                "请稍后在工作台任务面板确认任务状态，或重试。")
    if evidence is None:
        # task row 不存在（含孤儿事件场景：agent_events 有记录但 tasks 行缺失）
        return ("物流后台任务**未确认创建成功**（任务行不存在）。"
                "请稍后在工作台任务面板确认任务状态，或重试。")
    if not evidence.get("events"):
        return ("物流后台任务**未确认创建成功**（任务记录或事件缺失）。"
                "请稍后在工作台任务面板确认任务状态，或重试。")
    return None


# ── T07 freshness gate（确定性运营查询预检）────────────────────────────────
# 在 LLM 调用前插入：识别"最新/今天/TopN 销量"类运营查询 → 检查业务日覆盖 → 数据不足时
# 直接触发 workflow 或返回结构化不可用，禁止 LLM 自由补数（workflow_task=null + 模拟数事故源）。
_FRESHNESS_GATE_SALES_RE = _re.compile(
    # 显式时间窗 + 销售意图
    r"(?:今天|今日|最新|本周|这周|最近[0-9一两三四五六七八九十]+天?).*?(?:卖|销量|销售|热销|top\s*\d|前\s*\d|排名)"
    # 纯销售排名短语（无时间也隐含"最近"）
    r"|(?:卖得最好|卖得最多|热销|热门|销量最高|销量最多|最畅销|最好卖)"
    r"|(?:前[0-9]+|top\s*[0-9]+).*?(?:销量|卖|热销)"
    r"|哪[些个].*?(?:卖得最好|卖得最多|销量最高|最畅销|最好卖)",
    _re.IGNORECASE | _re.DOTALL,
)
# 明确拒绝刷新的否定短语 → 不触发 gate（用户想用现有数据答）
_FRESHNESS_GATE_SKIP_RE = _re.compile(
    r"(?:不用|不要|无需|先别).{0,8}(?:刷新|更新|同步)|就用现在的|先告诉我|不用等",
)
# WS-119: 库存类批量/榜单/约束查询（排行、可售、缺货、积压…）需 freshness gate 路由。
# 只匹配「批量/排序/数量约束」意图；单 SKU 实时问题由编码排除（见 _SKU_OR_ORDER_CODE_RE）。
_FRESHNESS_GATE_STOCK_RE = _re.compile(
    r"(?:库存|可售|缺货|断货|积压|备货).*?(?:排行|排名|榜|最多|最高|最大|最低|最少|多少|够不够|不够|缺口|清单|哪[些个]|top\s*\d|前\s*\d)"
    r"|(?:哪[些个]|多少).*?(?:库存|可售|缺货|断货|积压)"
    r"|(?:库存|可售|积压).*?(?:排行|排名|top\s*\d|前\s*\d)",
    _re.IGNORECASE | _re.DOTALL,
)
# WS-119: 物流类批量/榜单查询（在途/卡单/滞留排行、汇总）需 freshness gate 路由。
# 单 SKU/单货单实时问题继续优先走 query_sku_live/query_order_live（编码排除 + 既有 order_live 路由）。
_FRESHNESS_GATE_LOGISTICS_RE = _re.compile(
    r"(?:在途|卡单|滞留|压货|物流|货单|发货).*?(?:排行|排名|榜|最多|最高|多少|总量|汇总|清单|哪[些个]|top\s*\d|前\s*\d)"
    r"|(?:哪[些个]|多少).*?(?:在途|卡单|滞留|货单)"
    r"|(?:卡单|滞留).*?(?:货单|批次|sku|SKU)",
    _re.IGNORECASE | _re.DOTALL,
)
# WS-119 验收③：带明确 SKU/货单编码的单点实时问题 → 不当批量榜单 gate，交既有 live 工具（防退化成旧缓存）。
# 复用 _extract_live_order_no 的编码形态（2+ 字母 + 5+ 位字母数字/连字符），避免误吞 "top5"/"SKU" 这类非编码词。
_SKU_OR_ORDER_CODE_RE = _re.compile(r"\b[A-Z]{2,}[A-Z0-9-]{5,}\b")
_FRESHNESS_TARGET_ISO_DATE_RE = _re.compile(r"\b(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})\b")
_FRESHNESS_TARGET_CN_DATE_RE = _re.compile(r"(20\d{2})年\s*(\d{1,2})月\s*(\d{1,2})日?")


def _extract_freshness_target_date(question: str) -> Optional[str]:
    """Extract an explicit target business date/window end_date from a question."""
    import datetime as _dt

    candidates = []
    for rx in (_FRESHNESS_TARGET_ISO_DATE_RE, _FRESHNESS_TARGET_CN_DATE_RE):
        for m in rx.finditer(question or ""):
            try:
                candidates.append(_dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3))))
            except ValueError:
                continue
    if not candidates:
        return None
    return max(candidates).isoformat()


def _detect_operational_domain(question: str) -> Optional[str]:
    """T07: 识别需要 freshness gate 的运营查询类型。

    返回值:
      'sales'      — 销量类查询，需要 freshness gate 路由
      'sales_skip' — 用户明确说不用刷新，但仍在问销量排名；gate 跳过但需注入陈旧提示
      'stock'      — 库存类批量/榜单/约束查询（WS-119）
      'logistics'  — 物流类批量/榜单查询（WS-119）
      None         — 无需 gate（非运营查询、明确说不刷、或带 SKU/货单编码的单点实时问题）
    """
    q = question or ""
    skip = bool(_FRESHNESS_GATE_SKIP_RE.search(q))
    sales = bool(_FRESHNESS_GATE_SALES_RE.search(q))
    if skip and sales:
        return "sales_skip"
    if skip:
        return None
    if sales:
        return "sales"
    # WS-119：库存/物流批量榜单查询接入同一 freshness gate。
    # 验收③：带明确 SKU/货单编码的单点实时问题不被批量 gate 捕获——交既有
    # query_sku_live / query_order_live（防退化成旧缓存）。
    has_code = bool(_SKU_OR_ORDER_CODE_RE.search(q.upper()))
    if not has_code:
        if _FRESHNESS_GATE_STOCK_RE.search(q):
            return "stock"
        if _FRESHNESS_GATE_LOGISTICS_RE.search(q):
            return "logistics"
    return None


def _freshness_gate_route(store: str, question: str, scope: Dict) -> Optional[Dict]:
    """T07: LLM 调用前确定性 freshness 路由。
    返回完整 chat response dict（直接返给调用方）；若数据已覆盖/无法匹配则返 None（继续走 LLM）。
    特例：返回 {"_stale_skip": True, "_stale_suffix": "..."} 时，调用方应继续走 LLM 并在
    LLM 回复后追加 _stale_suffix（确定性陈旧警示，避免依赖 LLM wording 导致 T07-2 flaky）。
    """
    from . import _provider as _prov
    domain = _detect_operational_domain(question)
    if not domain:
        return None
    target_date = _extract_freshness_target_date(question)

    # sales_skip: 用户明确说不刷新但仍问销量排名。不拦截 LLM，但检查数据新鲜度，
    # 若数据陈旧则返回确定性陈旧后缀供调用方追加（代码级注入，不依赖 LLM wording）。
    if domain == "sales_skip":
        freshness = _data.check_freshness_coverage(store, "sales", target_date)
        latest = freshness.get("latest_date") or ""
        target = freshness.get("target_date") or ""
        if freshness.get("covered"):
            suffix = (
                "\n\n（提示：按你的要求本轮没有刷新，直接使用当前销量数据"
                + (f"，最新到 {latest}" if latest else "")
                + "；noon 销量同步可能滞后，结果偏保守。）"
            )
            return {"_stale_skip": True, "_stale_suffix": suffix}
        target_s = f"目标日期 {target} 暂未覆盖" if target_date else "未更新到今天"
        suffix = (
            f"\n\n（⚠️ 提示：当前销量数据{target_s}"
            + (f"，最新到 {latest}" if latest else "")
            + "，如需最新数据请随时刷新。）"
        )
        return {"_stale_skip": True, "_stale_suffix": suffix}

    freshness = _data.check_freshness_coverage(store, domain, target_date)
    if freshness.get("covered"):
        return None  # 数据新鲜 → 继续走 LLM/既有确定性路由（用最新业务日算）

    # WS-119：文案按域走，库存/物流不串"销量"措辞。
    domain_label = {"sales": "销量", "stock": "库存", "logistics": "物流"}.get(domain, domain)
    action = freshness.get("action") or "unavailable"
    latest = freshness.get("latest_date") or ""
    target = freshness.get("target_date") or ""
    wf = freshness.get("workflow")
    when_s = f"最新到 {latest}" if latest else "暂无数据"

    if action == "run_workflow" and wf:
        tool_result = _exec_tool("run_workflow", {"workflow": wf, "followup_prompt": question}, user=scope)
        workflow_task = None
        if isinstance(tool_result, dict) and tool_result.get("ok"):
            workflow_task = {
                "task_id": tool_result["task_id"],
                "workflow": tool_result["workflow"],
                "label": tool_result["label"],
                "total_steps": tool_result["total_steps"],
                "affected_modules": tool_result["affected_modules"],
                "followup_prompt": tool_result.get("followup_prompt"),
            }
            reply = (
                f"{domain_label}数据{when_s}，目标日期 {target} 暂未覆盖，"
                f"已触发更新（{wf}）。跑完后我会接着告诉你。"
            )
        else:
            err = (tool_result or {}).get("error") or ""
            reply = f"{domain_label}数据{when_s}，目标日期 {target} 暂未覆盖，更新触发失败：{err}"
        return {
            "reply": reply, "clean_reply": reply,
            "references": _dedup_refs((tool_result or {}).get("references", [])),
            "action_id": None, "tools_used": ["run_workflow"], "tag": "执行",
            "workflow_task": workflow_task,
            "provider": _prov.get_provider(), "confidence": 1.0,
            "judge_method": "freshness_gate", "freshness_gate": freshness,
            "hallucination_warnings": None,
        }

    if action == "upload_csv":
        csv_hint = freshness.get("csv_hint") or {}
        reply = (
            f"{domain_label}数据{when_s}，目标日期 {target} 暂未覆盖，无法自动刷新此部分。\n\n"
            f"👉 请到{csv_hint.get('where', '紫鸟 noon 后台')}导出 CSV，"
            f"文件名形如 `{csv_hint.get('csv_pattern', 'sales_noon_*.csv')}`，"
            "拖到工作台 📤 上传区。上传后我会接着告诉你。"
        )
        return {
            "reply": reply, "clean_reply": reply, "references": [],
            "action_id": None, "tools_used": [], "tag": "信息",
            "workflow_task": None,
            "provider": _prov.get_provider(), "confidence": 1.0,
            "judge_method": "freshness_gate", "freshness_gate": freshness,
            "hallucination_warnings": None,
        }

    # fallback: 数据不足，无 workflow 可跑
    reply = f"数据不足：{domain_label}{when_s}，无法提供 {target} 的查询结果。"
    return {
        "reply": reply, "clean_reply": reply, "references": [],
        "action_id": None, "tools_used": [], "tag": "信息",
        "workflow_task": None,
        "provider": _prov.get_provider(), "confidence": 1.0,
        "judge_method": "freshness_gate", "freshness_gate": freshness,
        "hallucination_warnings": None,
    }


def _execute_workflow_route(
    direct_workflow: Dict[str, str],
    question: str,
    scope: Dict,
    judge_method: str = "deterministic_workflow_router",
) -> Dict:
    """真实工作流执行路由（确定性）：调 run_workflow、构造 workflow_tasks + 三态受理回执。

    WS-145 肯定执行 + WS-159 库存刷新确认门共用此路由 —— 一处真实触发逻辑，避免确认轮另写
    一份「执行」分支导致接线/回执口径漂移。绝不返回「已触发/已完成」假证据，task_id 来自真实
    tool_result / 既有运行实例。
    """
    tool_args = {"workflow": direct_workflow["workflow"], "followup_prompt": question}
    tool_result = _exec_tool("run_workflow", tool_args, user=scope)
    workflow_tasks = []
    if isinstance(tool_result, dict) and tool_result.get("ok"):
        task_id = tool_result["task_id"]
        workflow_tasks.append({
            "ok": True,
            "task_id": task_id,
            "workflow": tool_result.get("workflow", direct_workflow["workflow"]),
            "label": tool_result.get("label", direct_workflow.get("label", direct_workflow["workflow"])),
            "total_steps": tool_result.get("total_steps", 0),
            "affected_modules": tool_result.get("affected_modules", []),
            "followup_prompt": tool_result.get("followup_prompt"),
        })
        # T21-SUB-2: 三态受理回执（已排队/已开始/已完成·失败），
        # 直接回答「是否已创建」并附 task_id/workflow/状态，禁止只说「已触发」。
        reply = _workflow_receipt_reply(
            task_id, tool_result["workflow"], direct_workflow["label"]
        )
        impact = _workflow_business_impact_reply(tool_result["workflow"], question)
        if impact:
            reply = f"{reply}\n\n{impact}"
        # T21-SUB-3 物流入口专项降级：用回读接口验证 durable 任务证据；
        # 任务表报错或事件缺失时降级回复，不返回假成功。
        if direct_workflow.get("workflow") == "wf3_logistics_v2":
            degrade_msg = _logistics_task_evidence_check(task_id)
            if degrade_msg:
                reply = degrade_msg
    elif (
        isinstance(tool_result, dict)
        and tool_result.get("action_type") == "denied"
        and "已有运行中实例" in (tool_result.get("reason") or "")
    ):
        existing = _current_workflow_task(direct_workflow["workflow"])
        if existing:
            workflow_tasks.append({
                "ok": True,
                "task_id": existing["task_id"],
                "workflow": existing["workflow"],
                "label": direct_workflow["label"],
                "total_steps": None,
                "affected_modules": [],
                "followup_prompt": question,
                "state": existing.get("state"),
                "already_running": True,
            })
            reply = (
                f"{direct_workflow['label']}已有运行中的后台任务 "
                f"`{existing['task_id']}`，我不重复触发。"
            )
        else:
            # DB not yet updated; parse ID from denial reason string
            extracted_id = _existing_workflow_task_id(tool_result)
            if extracted_id:
                workflow = direct_workflow["workflow"]
                label, total_steps, affected = _workflow_registry_summary(
                    workflow, direct_workflow["label"]
                )
                workflow_tasks.append({
                    "ok": True,
                    "task_id": extracted_id,
                    "workflow": workflow,
                    "label": label,
                    "total_steps": total_steps,
                    "affected_modules": affected,
                    "followup_prompt": question,
                })
                reply = (
                    f"{direct_workflow['label']}已有同类后台任务在运行，未新建重复任务。\n"
                    f"任务 ID：{extracted_id}｜workflow：{workflow}｜当前状态：运行中或排队。\n"
                    "请在工作台任务面板查看进度；任务结束后如仍需刷新，可以再重试。"
                )
            else:
                reply = tool_result.get("reason") or "工作流触发失败。"
    else:
        existing_task_id = _existing_workflow_task_id(tool_result or {})
        if existing_task_id:
            workflow = direct_workflow["workflow"]
            label, total_steps, affected = _workflow_registry_summary(
                workflow, direct_workflow["label"]
            )
            workflow_tasks.append({
                "ok": True,
                "task_id": existing_task_id,
                "workflow": workflow,
                "label": label,
                "total_steps": total_steps,
                "affected_modules": affected,
                "followup_prompt": question,
            })
            reply = (
                f"{direct_workflow['label']}已有同类后台任务在运行，未新建重复任务。\n"
                f"任务 ID：{existing_task_id}｜workflow：{workflow}｜当前状态：运行中或排队。\n"
                "请在工作台任务面板查看进度；任务结束后如仍需刷新，可以再重试。"
            )
        else:
            workflow_tasks.append({
                "ok": False,
                "workflow": direct_workflow["workflow"],
                "label": direct_workflow.get("label", direct_workflow["workflow"]),
                "error": (tool_result or {}).get("error") or "触发失败",
                "task_id": None,
            })
            reason = (
                (tool_result or {}).get("message")
                or (tool_result or {}).get("error")
                or (tool_result or {}).get("reason")
            )
            if reason:
                reply = reason
            elif direct_workflow.get("workflow") == "wf3_logistics_v2":
                reply = ("物流后台任务**未确认创建成功**（工作流触发失败）。"
                         "请稍后在工作台任务面板确认任务状态，或重试。")
            else:
                reply = "本轮没有创建后台任务：工作流触发失败，请稍后重试。"
            # WS-145 自动补调策略:确定性路由这一次触发即「自动补调一次」。
            # 失败后 policy 判定转 plan→confirm（不无限重试），追加下一步 + 需确认，
            # 绝不返回「已触发/已完成」假证据。
            from . import _execution_intent_gate as _intent_gate
            if (
                _intent_gate.decide_recovery(_intent_gate.RiskTier.LOW_AUTO, 1)
                == _intent_gate.RecoveryAction.PLAN_CONFIRM
            ):
                reply = reply.rstrip() + (
                    "\n\n下一步:这步自动补调一次仍未成功，我不再自动重复触发——"
                    "回「确认」我再试一次，或回「取消」改用上传/手动核对。"
                )
    from . import _provider
    return {
        "reply": reply,
        "clean_reply": reply,
        "references": _dedup_refs((tool_result or {}).get("references", [])),
        "action_id": None,
        "tools_used": ["run_workflow"],
        "tag": "执行",
        "workflow_tasks": workflow_tasks,
        "provider": _provider.get_provider(),
        "confidence": 1.0,
        "judge_method": judge_method,
        "hallucination_warnings": None,
    }


def _msg_text(m) -> str:
    """从一条 message 取纯文本（content 可能是 blocks 列表）。"""
    if not isinstance(m, dict):
        return ""
    c = m.get("content")
    if isinstance(c, list):
        return " ".join(b.get("text", "") for b in c if isinstance(b, dict))
    return c or ""


def _inventory_refresh_feasibility(scope: Dict):
    """库存刷新「可执行性」稳口径判断（WS-159）。

    不仅判低风险，还要能锁定范围 + 无明显执行前阻断：
      - 缺店铺范围（不知道刷哪个店铺）→ 不可执行；
      - 已有正在运行的库存刷新任务（冲突）→ 不可执行（不重复触发）。
    其余视为可执行。返回 (ok: bool, reason: str, next_step: str)。
    """
    store = (scope or {}).get("store")
    if not store:
        return (
            False,
            "缺少店铺范围（不知道要刷哪个店铺的库存）",
            "请先在工作台选好店铺（如 KSA），再让我刷新",
        )
    existing = _current_workflow_task("wf1_stock_v2")
    if existing:
        return (
            False,
            f"{store} 已有一个正在运行的库存刷新任务（task `{existing['task_id']}`）",
            "等当前任务跑完再刷，或在任务面板查看进度",
        )
    return (True, "", "")


def _pending_inventory_refresh_inquiry(messages: List[Dict]) -> Optional[str]:
    """从消息历史结构性推出「上一轮存在可执行库存刷新提议」(pending)。

    判据（只看紧接上一轮，pending 只对下一轮有效）：
      messages[-3] 是 user 且为询问式库存刷新请求，
      messages[-2] 是 assistant 且含本门的 PROPOSAL_MARKER（=确实提过可执行刷新）。
    满足则返回那条询问文本（用作 followup_prompt）；否则 None。
    换题后 messages[-3] 不再是询问句 → 自然返回 None（pending 失效）。
    """
    from . import _inventory_refresh_gate as _inv_gate
    if not messages or len(messages) < 3:
        return None
    prev_assistant = messages[-2]
    prev_user = messages[-3]
    if not isinstance(prev_assistant, dict) or prev_assistant.get("role") != "assistant":
        return None
    if not isinstance(prev_user, dict) or prev_user.get("role") != "user":
        return None
    inquiry = _msg_text(prev_user)
    if not _inv_gate.is_inventory_refresh_inquiry(inquiry):
        return None
    if _inv_gate.PROPOSAL_MARKER not in _msg_text(prev_assistant):
        return None
    return inquiry


def _inventory_refresh_no_task_result(reply: str, judge_method: str) -> Dict:
    from . import _provider
    return {
        "reply": reply,
        "clean_reply": reply,
        "references": [],
        "action_id": None,
        "tools_used": [],
        "tag": "确认",
        "workflow_task": None,
        "workflow_tasks": [],
        "provider": _provider.get_provider(),
        "confidence": 1.0,
        "judge_method": judge_method,
        "hallucination_warnings": None,
    }


def _inventory_refresh_confirm_gate(
    messages: List[Dict], question: str, scope: Dict
) -> Optional[Dict]:
    """WS-159 库存刷新询问式确认门。返回 chat 结果 dict 即短路；返回 None 表示不介入。

    三类轮次：
      1) 裸确认轮：有 pending → 执行一次真实 wf1_stock_v2；无 pending → 要求说明要执行什么。
      2) 取消轮：有 pending → 作废、不执行；无 pending → 交既有流程（普通否定句）。
      3) 询问轮：可执行 → 提议 + 反问确认（挂 pending）；不可/缺信息 → 说明缺口（不挂 pending）。
    """
    from . import _inventory_refresh_gate as _inv_gate
    q = question or ""

    # 1) 裸确认轮
    if _inv_gate.is_confirmation(q):
        inquiry = _pending_inventory_refresh_inquiry(messages)
        if inquiry is not None:
            ok, reason, next_step = _inventory_refresh_feasibility(scope)
            if not ok:
                return _inventory_refresh_no_task_result(
                    _inv_gate.pending_now_infeasible_reply(reason, next_step),
                    "inventory_refresh_confirm_now_infeasible",
                )
            # 消费 pending：执行一次真实库存刷新（复用统一执行路由）。
            return _execute_workflow_route(
                {"workflow": "wf1_stock_v2", "label": "库存刷新"},
                inquiry,
                scope,
                judge_method="inventory_refresh_confirm_consumed",
            )
        return _inventory_refresh_no_task_result(
            _inv_gate.bare_confirm_no_pending_reply(),
            "inventory_refresh_no_pending",
        )

    # 2) 取消轮
    if _inv_gate.is_cancellation(q):
        if _pending_inventory_refresh_inquiry(messages) is not None:
            return _inventory_refresh_no_task_result(
                _inv_gate.cancelled_reply(), "inventory_refresh_cancelled"
            )
        return None  # 无 pending 的否定句 → 交既有 WS-145 流程

    # 3) 询问轮（turn 1）
    if _inv_gate.is_inventory_refresh_inquiry(q):
        ok, reason, next_step = _inventory_refresh_feasibility(scope)
        if ok:
            store = (scope or {}).get("store")
            return _inventory_refresh_no_task_result(
                _inv_gate.proposal_reply(store), "inventory_refresh_proposed"
            )
        return _inventory_refresh_no_task_result(
            _inv_gate.infeasible_reply(reason, next_step),
            "inventory_refresh_infeasible",
        )

    return None


def chat(messages: List[Dict], scope: Dict) -> Dict:
    """
    messages: [{role: 'user'|'assistant', content: '...'}]
    scope: {store, current_user, current_role, tenant_id, user_id, ...}
    返回: {reply, clean_reply, references, action_id, tag, workflow_tasks, tools_used, provider, confidence}

    走 _provider 抽象层，通过 LLM_PROVIDER env 切换 anthropic / qwen / deepseek / doubao。
    """
    from . import _provider

    # 把 scope.tenant_id 注入 contextvars，让所有 tool 函数（同线程）能拿到
    _chat_tenant.set(scope.get("tenant_id") or 1)
    _chat_scope.set(scope)
    _last_replenishment_stock_status.set(None)
    _last_sku_rate_stats.set(None)
    # 同时设给 data 层（PG RLS 用）
    _data.set_current_tenant(scope.get("tenant_id") or 1)

    question = messages[-1].get("content") if messages else ""
    if isinstance(question, list):  # content 可能是 blocks
        question = " ".join(b.get("text", "") for b in question if isinstance(b, dict))
    _chat_question.set(question or "")

    # WS-145 肯定执行意图门:本轮句式语气 + 风险分层一次求出，注入 contextvar，
    # 供 _exec_tool 在非执行语气下拒绝 run_workflow（LLM 不许绕）。
    from . import _execution_intent_gate as _intent_gate
    _intent_decision = _intent_gate.evaluate(question or "")
    _chat_intent.set(_intent_decision)

    # WS-150: 工作台不支持主动发飞书/通知群（确定性拒绝，不进 confirm-first）
    if _intent_decision.unsupported_feishu_notify:
        reply = _intent_gate.unsupported_feishu_notify_reply()
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": [],
            "action_id": None,
            "tools_used": [],
            "tag": "拒绝",
            "workflow_task": None,
            "workflow_tasks": [],
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "execution_intent_gate_unsupported_feishu_notify",
            "hallucination_warnings": None,
        }

    # 高风险动作（外部通知/交易·采购·订单/不可回滚/跨店批量覆盖）即使肯定句也不自动执行:
    # 先 confirm，不自动补调。确定性短路，绝不落任务。
    if _intent_decision.needs_confirm_first:
        reply = _intent_gate.confirm_first_reply(question or "")
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": [],
            "action_id": None,
            "tools_used": [],
            "tag": "确认",
            "workflow_task": None,
            "workflow_tasks": [],
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "execution_intent_gate_confirm_first",
            "hallucination_warnings": None,
        }

    # WS-159 库存刷新询问式确认门:跨轮 pending 解锁。
    #   询问轮（能不能帮我刷新库存?）→ 提议 + 反问，不落任务;
    #   裸确认轮（好/可以/确认）且上一轮有提议 → 只执行一次真实 wf1_stock_v2;
    #   取消/换题/模糊/无 pending 裸确认 → 不执行。
    # 高风险询问不入此门（由上方 confirm-first / _exec_tool 兜），一句「好」不解锁高风险。
    inv_refresh = _inventory_refresh_confirm_gate(messages, question, scope)
    if inv_refresh is not None:
        return inv_refresh

    direct_export = _deterministic_export_request(question)
    if direct_export:
        store = scope.get("store") or "KSA"
        tool_args = {
            "view": direct_export["view"],
            "store": store,
            "filter_desc": direct_export["filter_desc"],
        }
        tool_result = _exec_tool("export_table", tool_args, user=scope)
        if isinstance(tool_result, dict) and tool_result.get("ok") and tool_result.get("download_url"):
            filename = tool_result.get("filename") or "export.xlsx"
            reply = (
                f"已生成 {tool_result.get('row_count', 0)} 行表格："
                f"[{filename}]({tool_result['download_url']})"
            )
        else:
            reply = (tool_result or {}).get("message") or (tool_result or {}).get("error") or "导出失败。"
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": [],
            "action_id": None,
            "tools_used": ["export_table"],
            "tag": "查询",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "deterministic_export_router",
            "hallucination_warnings": None,
        }

    direct_readonly = _deterministic_readonly_request(question)
    if direct_readonly:
        store = (scope.get("store") or "KSA").upper()
        tool_args = {"store": store}
        tool_result = _exec_tool(direct_readonly["tool"], tool_args, user=scope)
        reply = _deterministic_readonly_reply(
            direct_readonly["intent"], tool_result or {}, store
        )
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": _dedup_refs((tool_result or {}).get("references", [])),
            "action_id": None,
            "tools_used": [direct_readonly["tool"]],
            "tag": "查询",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "deterministic_readonly_router",
            "hallucination_warnings": None,
        }

    # T37 round-15（Luke 指令①）：库存刷新副作用动作，用户明确拒绝（任意语序）→
    # 确定性回复，绝不调 run_workflow，绝不伪造任务号/已启动声明。
    stock_refusal_reply = _stock_refresh_refusal_reply(question)
    if stock_refusal_reply:
        return {
            "reply": stock_refusal_reply,
            "clean_reply": stock_refusal_reply,
            "references": [],
            "action_id": None,
            "tools_used": [],
            "tag": "执行",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "deterministic_stock_refusal_router",
            "hallucination_warnings": None,
        }

    if _deterministic_erp_refresh_time_request(question):
        store = (scope.get("store") or "KSA").upper()
        reply = _format_erp_refresh_time_reply(store, _data.get_data_health(store))
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": [],
            "action_id": None,
            "tools_used": [],
            "tag": "查询",
            "workflow_task": None,
            "workflow_tasks": [],
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "deterministic_erp_refresh_time_router",
            "hallucination_warnings": None,
        }

    if _deterministic_data_freshness_request(question):
        store = scope.get("store") or "KSA"
        tool_result = _exec_tool("data_health_check", {"store": store}, user=scope)
        reply = _format_data_freshness_reply(store, tool_result)
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": _dedup_refs((tool_result or {}).get("references", [])),
            "action_id": None,
            "tools_used": ["data_health_check"],
            "tag": "查询",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "deterministic_data_freshness_router",
            "hallucination_warnings": None,
        }

    # WS-98 Round-2：非执行语气的刷新/重算意图（询问/假设/只问影响面）确定性短路，
    # 交回 WS-145 结构门的干净解释，绝不落 LLM。否则 LLM 会去试 run_workflow——
    # 虽被 _exec_tool 拦下不落任务，却会污染 tools_used 且触发 _safety 假活 banner，
    # 把「能不能帮我刷新…?」误渲染成带警告/「启动失败」的回复（验门人 Round-2 红队洞）。
    # NEGATED 不在此短路：它常是「不用刷新，但告诉我哪些要补」这类仍要数据答案的句子，
    # 留给 LLM 给陈旧警示 + 答案（smoke「用户拒绝刷新」）。
    if (
        _intent_decision.has_refresh_trigger
        and _intent_decision.mood in (
            _intent_gate.IntentMood.INTERROGATIVE,
            _intent_gate.IntentMood.HYPOTHETICAL,
            _intent_gate.IntentMood.IMPACT_QUERY,
        )
        and not _deterministic_sku_metric_request(question)
    ):
        reply = _intent_gate.explain_reply(_intent_decision.mood, question)
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": [],
            "action_id": None,
            "tools_used": [],
            "tag": "查询",
            "workflow_task": None,
            "workflow_tasks": [],
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "execution_intent_gate_explain_non_executory",
            "hallucination_warnings": None,
        }

    direct_workflows = _deterministic_multi_workflow_request(question)
    if direct_workflows:
        workflow_tasks = []
        reply_parts = []
        for direct_workflow in direct_workflows:
            workflow = direct_workflow["workflow"]
            label = direct_workflow["label"]
            tool_args = {"workflow": workflow, "followup_prompt": question}
            tool_result = _exec_tool("run_workflow", tool_args, user=scope)
            if isinstance(tool_result, dict) and tool_result.get("ok"):
                task_id = tool_result["task_id"]
                workflow_tasks.append({
                    "ok": True,
                    "task_id": task_id,
                    "workflow": tool_result.get("workflow", workflow),
                    "label": tool_result.get("label", label),
                    "total_steps": tool_result.get("total_steps", 0),
                    "affected_modules": tool_result.get("affected_modules", []),
                    "followup_prompt": tool_result.get("followup_prompt"),
                })
                reply_parts.append(_workflow_receipt_reply(
                    task_id, tool_result.get("workflow", workflow), label
                ))
                continue

            existing_task_id = None
            if (
                isinstance(tool_result, dict)
                and tool_result.get("action_type") == "denied"
                and "已有运行中实例" in (tool_result.get("reason") or "")
            ):
                existing = _current_workflow_task(workflow)
                if existing:
                    existing_task_id = existing["task_id"]
                    workflow_tasks.append({
                        "ok": True,
                        "task_id": existing_task_id,
                        "workflow": existing["workflow"],
                        "label": label,
                        "total_steps": None,
                        "affected_modules": [],
                        "followup_prompt": question,
                        "state": existing.get("state"),
                        "already_running": True,
                    })
                    reply_parts.append(
                        f"{label}（{workflow}）已有运行中的后台任务 `{existing_task_id}`，"
                        "本轮未新建重复任务。请在工作台任务面板查看进度。"
                    )
                    continue
                existing_task_id = _existing_workflow_task_id(tool_result)

            if existing_task_id:
                label2, total_steps, affected = _workflow_registry_summary(workflow, label)
                workflow_tasks.append({
                    "ok": True,
                    "task_id": existing_task_id,
                    "workflow": workflow,
                    "label": label2,
                    "total_steps": total_steps,
                    "affected_modules": affected,
                    "followup_prompt": question,
                    "already_running": True,
                })
                reply_parts.append(
                    f"{label2}（{workflow}）已有同类后台任务在运行，未新建重复任务。\n"
                    f"任务 ID：{existing_task_id}｜当前状态：运行中或排队。\n"
                    "请在工作台任务面板查看进度；任务结束后如仍需刷新，可以再重试。"
                )
                continue

            reason = (
                (tool_result or {}).get("message")
                or (tool_result or {}).get("error")
                or (tool_result or {}).get("reason")
                or "触发失败"
            )
            workflow_tasks.append({
                "ok": False,
                "workflow": workflow,
                "label": label,
                "error": reason,
                "task_id": None,
            })
            reply_parts.append(f"{label}（{workflow}）启动失败：{reason}。")

        return {
            "reply": "\n\n".join(reply_parts),
            "clean_reply": "\n\n".join(reply_parts),
            "references": [],
            "action_id": None,
            "tools_used": ["run_workflow"],
            "tag": "执行",
            "workflow_tasks": workflow_tasks,
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "deterministic_multi_workflow_router",
            "hallucination_warnings": None,
        }

    direct_workflow = _deterministic_workflow_request(question)
    if direct_workflow:
        return _execute_workflow_route(direct_workflow, question, scope)

    direct_sales_topn_n = _deterministic_product_sales_topn_request(question)
    if direct_sales_topn_n is not None:
        store = scope.get("store") or "KSA"
        tool_result = _exec_tool(
            "list_products",
            {"store": store, "listing": "all", "limit": direct_sales_topn_n},
            user=scope,
        )
        reply = _format_product_sales_topn_reply(store, tool_result)
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": _dedup_refs((tool_result or {}).get("references", [])),
            "action_id": None,
            "tools_used": ["list_products"],
            "tag": "查询",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0 if not (tool_result or {}).get("error") else 0.8,
            "judge_method": "deterministic_product_sales_topn_router",
            "hallucination_warnings": None,
        }

    if _provider.get_provider() != "smoke" and _deterministic_products_count_request(question):
        store = scope.get("store") or "KSA"
        tool_result = _exec_tool("list_products", {"store": store, "listing": "all", "limit": 0}, user=scope)
        reply = _format_products_count_reply(store, tool_result)
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": _dedup_refs((tool_result or {}).get("references", [])),
            "action_id": None,
            "tools_used": ["list_products"],
            "tag": "查询",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "deterministic_products_count_router",
            "hallucination_warnings": None,
        }

    if _deterministic_scope_overview_request(question):
        store = scope.get("store") or "KSA"
        tool_result = _exec_tool("scope_overview", {"store": store}, user=scope)
        reply = _format_scope_overview_reply(store, tool_result)
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": _dedup_refs((tool_result or {}).get("references", [])),
            "action_id": None,
            "tools_used": ["scope_overview"],
            "tag": "查询",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "deterministic_scope_overview_router",
            "hallucination_warnings": None,
        }

    direct_replenishment_sku = _deterministic_replenishment_sku_request(question)
    if direct_replenishment_sku:
        store = scope.get("store") or "KSA"
        tool_result = _exec_tool(
            "query_replenishment_sku",
            {"sku": direct_replenishment_sku, "store": store},
            user=scope,
        )
        from . import replenishment_evidence as _rep
        reply = _rep.format_replenishment_sku_reply(tool_result)
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": _dedup_refs((tool_result or {}).get("references", [])),
            "action_id": None,
            "tools_used": ["query_replenishment_sku"],
            "tag": "查询",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0 if (tool_result or {}).get("ok") else 0.9,
            "judge_method": "deterministic_replenishment_sku_router",
            "hallucination_warnings": None,
        }

    direct_replenishment_limit = _deterministic_replenishment_list_request(question)
    if direct_replenishment_limit is not None:
        store = scope.get("store") or "KSA"
        tool_result = _exec_tool(
            "compute_replenishment",
            {"store": store, "limit": direct_replenishment_limit},
            user=scope,
        )
        reply = _format_replenishment_list_reply(store, tool_result)
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": _dedup_refs((tool_result or {}).get("references", [])),
            "action_id": None,
            "tools_used": ["compute_replenishment"],
            "tag": "查询",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0 if not (tool_result or {}).get("fail_closed") else 0.9,
            "judge_method": "deterministic_replenishment_list_router",
            "hallucination_warnings": None,
        }

    direct_stock_split_sku = _deterministic_stock_split_request(question)
    if direct_stock_split_sku:
        store = scope.get("store") or "KSA"
        tool_result = _exec_tool("query_stock_split", {"sku": direct_stock_split_sku, "store": store}, user=scope)
        reply = _format_stock_split_reply(direct_stock_split_sku, tool_result)
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": _dedup_refs((tool_result or {}).get("references", [])),
            "action_id": None,
            "tools_used": ["query_stock_split"],
            "tag": "查询",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "deterministic_stock_split_router",
            "hallucination_warnings": None,
        }

    direct_sku_metric = _deterministic_sku_metric_request(question)
    if direct_sku_metric:
        store = scope.get("store") or "KSA"
        tool_result = _exec_tool(
            "query_sku",
            {"skus": [direct_sku_metric], "store": store},
            user=scope,
        )
        reply = _format_sku_metric_reply(direct_sku_metric, tool_result)
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": _dedup_refs((tool_result or {}).get("references", [])),
            "action_id": None,
            "tools_used": ["query_sku"],
            "tag": "查询",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "deterministic_sku_metric_router",
            "hallucination_warnings": None,
        }

    direct_order_no = _extract_live_order_no(question)
    if direct_order_no:
        tool_result = _exec_tool("query_order_live", {"order_no": direct_order_no}, user=scope)
        reply = _format_order_live_reply(direct_order_no, tool_result)
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": _dedup_refs((tool_result or {}).get("references", [])),
            "action_id": None,
            "tools_used": ["query_order_live"],
            "tag": "查询",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "deterministic_order_live_router",
            "hallucination_warnings": None,
        }

    # T15 — 总库存 TopN 确定性路由（WS-139）
    direct_topn_n = _deterministic_total_stock_topn_request(question)
    if direct_topn_n is not None:
        store = scope.get("store") or "KSA"
        tool_result = _exec_tool("total_stock_topn", {"store": store, "n": direct_topn_n}, user=scope)
        reply = _format_total_stock_topn_reply(store, tool_result)
        return {
            "reply": reply,
            "clean_reply": reply,
            "references": _dedup_refs((tool_result or {}).get("references", [])),
            "action_id": None,
            "tools_used": ["total_stock_topn"],
            "tag": "查询",
            "workflow_task": None,
            "provider": _provider.get_provider(),
            "confidence": 1.0,
            "judge_method": "deterministic_total_stock_topn_router",
            "hallucination_warnings": None,
        }

    if r := _handle_icr_chat(question, _provider.get_provider()):
        return r

    # T07: freshness gate — 运营查询（TopN 销量等）在 LLM 前先检业务日覆盖
    store = scope.get("store", "KSA")
    gate_result = _freshness_gate_route(store, question, scope)
    _t07_stale_suffix = ""
    if gate_result is not None:
        if gate_result.get("_stale_skip"):
            # sales_skip 场景：继续走 LLM，事后追加确定性陈旧提示
            _t07_stale_suffix = gate_result.get("_stale_suffix", "")
        else:
            return gate_result

    sys_text = SYSTEM_PROMPT.format(scope=json.dumps(scope, ensure_ascii=False))
    result = _provider.chat_with_tools(
        messages=_clean_history(messages),   # 清掉历史里残留的 banner，断自激
        system=sys_text,
        tools=TOOLS,
        tool_funcs=TOOL_FUNCS,
        scope=scope,
    )
    clean_reply  = result["reply"]           # LLM 原文（无 banner）— 用于持久化 + 喂未来历史
    tool_log     = result["tool_log"]
    refs_collected = result["refs_collected"]
    workflow_tasks = result.get("workflow_tasks", [])
    tools_used     = [t["name"] for t in tool_log]

    # T07-2 sales_skip: 确定性陈旧后缀（代码级注入，不依赖 LLM wording）
    if _t07_stale_suffix:
        clean_reply += _t07_stale_suffix

    # Layer 3 hallucinate 后处理（上移自 api.py — 一处产生 warnings，既喂 confidence 又 sanitize）
    # final_text = 展示版（可能带 banner）；clean_reply = 持久化版（无 banner，防历史自激）
    from . import _safety
    final_text, hallu_warnings = _safety.sanitize_reply(clean_reply, tools_used, tool_log=tool_log, question=question)
    clean_reply = _strip_safety_banner(final_text)

    # WS-117 采购议价率口径生产接线（deterministic verifier，非 prompt）
    from hipop.rules.procurement_rate import check_procurement_rate_reply as _check_procurement_rate
    _procurement_warns = _check_procurement_rate(clean_reply)
    if _procurement_warns:
        hallu_warnings = list(hallu_warnings or []) + _procurement_warns
        if not final_text.startswith("⚠️"):
            _proc_banner = (
                "⚠️ **系统检测到采购议价率口径可能有误**：\n"
                + "\n".join(f"- {w}" for w in _procurement_warns)
                + "\n\n---\n\n"
            )
            final_text = _proc_banner + final_text

    final_text = _maybe_append_stock_readiness_warning(final_text)
    clean_reply = _maybe_append_stock_readiness_warning(clean_reply)
    final_text = _ensure_export_download_link(final_text, tool_log)
    clean_reply = _ensure_export_download_link(clean_reply, tool_log)
    final_text = _maybe_inject_missing_rates(final_text, question)
    clean_reply = _maybe_inject_missing_rates(clean_reply, question)
    final_text = _maybe_append_oldest_data_health_date(final_text, question, tools_used, scope)
    clean_reply = _maybe_append_oldest_data_health_date(clean_reply, question, tools_used, scope)
    final_text = _maybe_append_order_lookup_negative_hint(final_text, question, tools_used)
    clean_reply = _maybe_append_order_lookup_negative_hint(clean_reply, question, tools_used)
    final_text = _maybe_append_navigation_url(final_text, tool_log)
    clean_reply = _maybe_append_navigation_url(clean_reply, tool_log)

    # judge + confidence 真逻辑（混合：启发式 + 低置信/destructive 触发 LLM judge）
    judge, confidence, judge_method = _compute_judge_confidence(
        question, final_text, tool_log, refs_collected, hallu_warnings)

    # 低置信自动在 reply 头部加提示（_safety 已加 banner 时不重复，避免双 banner）
    if confidence < 0.6 and not hallu_warnings:
        final_text = (
            f"⚠️ 我对这个回答的置信度较低（{int(confidence*100)}%），"
            "建议你核实关键数字，或换个更明确的问法。\n\n---\n\n"
        ) + final_text

    # WS-26: 撞限（做不到/超范围）回复确定性补一句『要记成需求吗』offer。
    # display 版 + 持久化版都补，保证下一轮用户回『记一下』时对话连贯。
    final_text  = _maybe_append_feedback_offer(final_text, tools_used)
    clean_reply = _maybe_append_feedback_offer(clean_reply, tools_used)

    # 写入 agent_actions（reference 系统）
    action_id = None
    if final_text and (refs_collected or tool_log):
        try:
            first_tool_args = _safety._normalize_args(tool_log[0].get("args") or {}) if tool_log else {}
            action_id = _data.write_agent_action(
                store=scope.get("store", "KSA"),
                module="chat",
                action_type="execute",
                subject=(first_tool_args.get("sku") or first_tool_args.get("order_no")) if tool_log else None,
                judge=judge,
                pill_text="执行" if _safety._is_substantive_action(tool_log) else ("查询" if tool_log else "信息"),
                pill="info",
                confidence=confidence,
                options=[],
                references=_dedup_refs(refs_collected),
                owner=scope.get("current_user", "Cherry"),
            )
        except Exception:
            pass

    return {
        "reply": final_text.strip() or "(无回复)",          # 展示版（带 banner）给前端当场看
        "clean_reply": (clean_reply or "").strip() or "(无回复)",  # 无 banner 版给持久化，防历史自激
        "references": _dedup_refs(refs_collected),
        "action_id": action_id,
        "tools_used": tools_used,
        "tag": ("hallucinate" if hallu_warnings else ("执行" if _safety._is_substantive_action(tool_log) else ("查询" if tool_log else None))),
        "workflow_tasks": workflow_tasks,
        "provider": _provider.get_provider(),
        "confidence": round(confidence, 2),
        "judge_method": judge_method,
        "hallucination_warnings": hallu_warnings or None,
}


def _dedup_refs(refs: List[Dict]) -> List[Dict]:
    seen = set()
    out = []
    for r in refs:
        k = (r.get("table"), r.get("where"))
        if k in seen: continue
        seen.add(k)
        out.append(r)
    return out
