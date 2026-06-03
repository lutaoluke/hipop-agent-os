"""Smoke: 销量「最后输出数据」导出字段契约（WS-20 / WS-14-6）— fail-then-pass。

钉死的是**输出层**口径：数据落库（wf2_sku / wf2_orders v2）不等于运营拿得到。
本 smoke 走真正的导出生产入口 `agent.tool_export_table(view="sales")`（UI 下载按钮
和 /api/export 都打这条路径），打开生成的 xlsx，逐字段断言需求声明的全字段都在、
且值确实来自 v2 的 wf2_sku / wf2_orders（按 tenant+entity 过滤）。

三种死法对应：
  · 接线缺失 —— 不读 DB、直接断 export 生产函数的输出，证明消费端真读到 v2 数据。
  · 占位假数据 —— 动态字段（订单量/订单号/退货率/取消率/最新成交价/总销售额/异常标记/
                  销量评级/预测）由 wf_sales_static_v2.merge_entity_v2 从 wf2_orders 真算，
                  smoke 逐项对真值，不接受空列或写死。
  · 越权串租户 —— 另一 tenant 同国别同 SKU 必须不出现在导出里。

fail-then-pass 证明：
  改动前（WS-20 之前）sales view 的导出列是一个英文子集
  （partner_sku/title/sales_*d/... 共 16 列），缺 国别/店铺名/订单量/订单号/售卖形式/
  最新成交价/退货率/取消率/sales_120d/总销售额/异常标记/销量评级/10·30天预测——
  本 smoke 断言「表头 == SALES_EXPORT_SPEC 全 26 列」会直接 FAIL；
  补齐导出口径后 PASS。

跑法：
  python3 tests/smoke_sales_export_fields.py
  （make test 会自动聚合本文件）
"""
import os
import sys
import json
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
sys.path.insert(0, REPO)

# 必须在 import hipop.server.data 之前设好 SQLite 路径 + 清掉 DB_URL，否则按 PG 跑。
os.environ["HIPOP_DB"] = tempfile.NamedTemporaryFile(suffix=".db", delete=False).name
os.environ.pop("DB_URL", None)

TID = 7                  # 主租户
OTHER_TID = 99           # 越权对照租户
ALIAS = "smoke_ksa"
STORE = "SA"             # _resolve_entity_alias: SA → country SA

_TMP = []


def _fresh_db(data):
    path = tempfile.NamedTemporaryFile(suffix=".db", delete=False).name
    _TMP.append(path)
    data.DB_PATH = path
    conn = data.conn()
    with open(os.path.join(REPO, "db", "schema_v2.sql"), encoding="utf-8") as f:
        sql = f.read()
    cut = sql.find("DO $$")          # PG RLS policy 之后 SQLite 跳过
    if cut != -1:
        sql = sql[:cut]
    for stmt in sql.split(";"):
        s = stmt.strip()
        if s:
            conn.execute(s)
    conn.commit()
    return conn


def _seed_entity(conn, tid, store_name):
    conn.execute(
        "INSERT INTO sales_entities "
        "(tenant_id, alias, country, platform, store_name, store_id, currency) "
        "VALUES (?,?,?,?,?,?,?)",
        (tid, ALIAS, "SA", "Noon", store_name, 999, "SAR"),
    )


def _seed_sku(conn, tid, partner_sku):
    """静态 ERP 字段；动态字段交给 merge_entity_v2 从 wf2_orders 算。"""
    conn.execute(
        "INSERT INTO wf2_sku "
        "(tenant_id, entity_alias, partner_sku, erp_sku_id, noon_sku, product_id, "
        " title, fulfillment, brand, cost_price, currency, is_listed, "
        " latest_price, avg_price, latest_profit_rate, "
        " sales_10d, sales_30d, sales_60d, sales_90d, sales_120d, sales_180d) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (tid, ALIAS, partner_sku, "ERP-1", "NSK-1", "PRD-1",
         "便携榨汁杯 A 款", "FBN", "HIPOP", 60.0, "SAR", 1,
         120.0, 110.0, 0.30,         # latest_price=ERP 120（noon 将覆盖为 100 → price_mismatch）
         2, 3, 4, 5, 6, 8),
    )


def _seed_orders(conn, tid):
    """两条 noon 订单：1 正常 + 1 退货。"""
    rows = [
        # item_nr, order_date, status, is_cancelled, is_return, seller_price, customer_paid
        ("ORD-A", "2026-05-30", "delivered", 0, 0, 100.0, 110.0),
        ("ORD-B", "2026-05-20", "Customer Initiated Returns", 0, 1, 100.0, 90.0),
    ]
    for it, od, st, ic, ir, sp, cp in rows:
        conn.execute(
            "INSERT INTO wf2_orders "
            "(tenant_id, entity_alias, partner_sku, noon_sku, item_nr, order_date, "
            " status, is_cancelled, is_return, seller_price, customer_paid, currency) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (tid, ALIAS, "PSK-1", "NSK-1", it, od, st, ic, ir, sp, cp, "SAR"),
        )


def _read_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = list(rows[0])
    data_rows = [dict(zip(header, r)) for r in rows[1:]]
    return header, data_rows


class _Checker:
    def __init__(self):
        self.failures = []

    def __call__(self, name, cond, detail=""):
        if cond:
            print(f"  ✓ {name}")
        else:
            self.failures.append(name)
            print(f"  ✗ {name} {detail}")


def _approx(a, b, tol=1e-6):
    return a is not None and b is not None and abs(float(a) - float(b)) <= tol


def run():
    from hipop.server import data, agent
    from hipop.workflows import wf_sales_static_v2

    data.set_current_tenant(TID)
    agent._chat_tenant.set(TID)

    conn = _fresh_db(data)
    _seed_entity(conn, TID, "SMOKE OFFICIAL STORE")
    _seed_sku(conn, TID, "PSK-1")
    _seed_orders(conn, TID)
    # 越权对照：另一 tenant 同国别 + 同名 SKU，导出绝不能带出来
    _seed_entity(conn, OTHER_TID, "OTHER STORE")
    _seed_sku(conn, OTHER_TID, "OTHER-1")
    conn.commit()

    # 跑真正的动态合并（把 wf2_orders 视角 merge 回 wf2_sku 的契约字段）
    n = wf_sales_static_v2.merge_entity_v2(TID, ALIAS, conn)
    conn.commit()
    conn.close()
    assert n >= 1, "merge_entity_v2 没处理任何行"

    # ── 走真正的导出生产入口 ──
    res = agent.tool_export_table(view="sales", store=STORE, listing="all")
    check = _Checker()
    check("export ok", res.get("ok") is True, f"got {res!r}")
    fpath = res.get("file_path")
    if not fpath or not os.path.exists(fpath):
        check("xlsx 文件生成", False, f"file_path={fpath!r}")
        return check.failures
    _TMP.append(fpath)

    header, drows = _read_xlsx(fpath)

    # ── 1. 表头 == 需求声明的全字段口径（字面量，独立于实现）──
    #    改动前 sales view 导出的是英文子集（partner_sku/title/sales_*d/...），
    #    缺这里多数列 → header 不等 → FAIL；补齐导出口径后 PASS。
    expected_headers = [
        "国别", "店铺名", "主SKU", "PSKU", "订单量", "订单号", "商品标题", "售卖形式",
        "商品最新售价", "平均售价", "最新成交价", "最新利润率", "退货率", "取消率",
        "最新出单日期", "近10天销量", "近30天销量", "近60天销量", "近90天销量",
        "近120天销量", "近180天销量", "总销售额", "异常标记", "销量评级",
        "10天预测", "30天预测",
    ]
    missing = [h for h in expected_headers if h not in (header or [])]
    check(f"表头覆盖全 {len(expected_headers)} 列输出字段口径",
          header == expected_headers,
          f"\n    missing={missing}\n    got   ={header}")

    # ── 2. 租户隔离：只出主租户那一行 ──
    check("导出仅含主租户 1 行（越权 SKU 不串）",
          res.get("row_count") == 1 and len(drows) == 1,
          f"row_count={res.get('row_count')} rows={len(drows)}")
    if not drows:
        return check.failures
    r = drows[0]
    check("PSKU 仅 PSK-1（非 OTHER-1）", r.get("PSKU") == "PSK-1", f"got {r.get('PSKU')!r}")

    # ── 3. 逐字段断言：静态来自 wf2_sku，动态来自 wf2_orders 合并 ──
    # 来源 entity（sales_entities，同 tenant+entity 过滤）
    check("国别==SA（来自 sales_entities）", r.get("国别") == "SA", f"got {r.get('国别')!r}")
    check("店铺名==SMOKE OFFICIAL STORE", r.get("店铺名") == "SMOKE OFFICIAL STORE",
          f"got {r.get('店铺名')!r}")
    # 来源 wf2_sku 静态
    check("主SKU==PRD-1", r.get("主SKU") == "PRD-1", f"got {r.get('主SKU')!r}")
    check("商品标题", r.get("商品标题") == "便携榨汁杯 A 款", f"got {r.get('商品标题')!r}")
    check("售卖形式==FBN", r.get("售卖形式") == "FBN", f"got {r.get('售卖形式')!r}")
    check("最新利润率==0.30", _approx(r.get("最新利润率"), 0.30), f"got {r.get('最新利润率')!r}")
    # 来源 wf2_orders（经 merge_entity_v2）
    check("订单量==2（wf2_orders 计数）", r.get("订单量") == 2, f"got {r.get('订单量')!r}")
    check("订单号==ORD-A,ORD-B（wf2_orders item_nr）",
          r.get("订单号") == "ORD-A,ORD-B", f"got {r.get('订单号')!r}")
    check("商品最新售价==100（noon 视角覆盖 ERP）", _approx(r.get("商品最新售价"), 100.0),
          f"got {r.get('商品最新售价')!r}")
    check("平均售价==100", _approx(r.get("平均售价"), 100.0), f"got {r.get('平均售价')!r}")
    check("最新成交价==110（最近一单 customer_paid）", _approx(r.get("最新成交价"), 110.0),
          f"got {r.get('最新成交价')!r}")
    check("退货率==0.5（1 退 / 2 有效）", _approx(r.get("退货率"), 0.5), f"got {r.get('退货率')!r}")
    check("取消率==0.0", _approx(r.get("取消率"), 0.0), f"got {r.get('取消率')!r}")
    check("最新出单日期==2026-05-30", r.get("最新出单日期") == "2026-05-30",
          f"got {r.get('最新出单日期')!r}")
    check("总销售额==200", _approx(r.get("总销售额"), 200.0), f"got {r.get('总销售额')!r}")
    check("异常标记==price_mismatch（noon 100 vs ERP 120）",
          r.get("异常标记") == "price_mismatch", f"got {r.get('异常标记')!r}")
    # 销量窗口（wf2_sku）
    for h, exp in (("近10天销量", 2), ("近30天销量", 3), ("近60天销量", 4),
                   ("近90天销量", 5), ("近120天销量", 6), ("近180天销量", 8)):
        check(f"{h}=={exp}", r.get(h) == exp, f"got {r.get(h)!r}")
    # 评级 / 预测（merge 算出，非空）
    check("销量评级 非空", bool(r.get("销量评级")), f"got {r.get('销量评级')!r}")
    check("10天预测 非空", r.get("10天预测") is not None, f"got {r.get('10天预测')!r}")
    check("30天预测 非空", r.get("30天预测") is not None, f"got {r.get('30天预测')!r}")

    return check.failures


if __name__ == "__main__":
    try:
        failures = run()
    finally:
        for p in _TMP + [os.environ.get("HIPOP_DB")]:
            try:
                if p:
                    os.unlink(p)
            except OSError:
                pass
    if failures:
        print(f"\n✗ {len(failures)} 项断言失败: {failures}")
        sys.exit(1)
    print("\n✓ 销量「最后输出数据」导出字段契约 smoke 全过")
    sys.exit(0)
